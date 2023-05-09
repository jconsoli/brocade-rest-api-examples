#!/usr/bin/python
# -*- coding: utf-8 -*-
# Copyright 2022, 2023 Jack Consoli.  All rights reserved.
#
# NOT BROADCOM SUPPORTED
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may also obtain a copy of the License at
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""
:mod:`certs_eval` - General report on certs

**Description**

    The purpose of this module is to illustrate how to read certificates, check the expiration dates, generate a CSR,
    delete certificates, and add certificates. For most customers, this will be done in a single pass of a script but
    since different organizations have different policies and methods for generating certificates, this module takes
    the following action types as input:

    * Read and evaluate certificates
    * Generate CSRs
    * Import (add) certificates
    * Delete certificates.

    To facilitate passing parameters via different passes, I used an Excel workbook. I'm assuming most people will use a
    different interface so I didn't spend a lot of time on error checking and other niceties. It's just good enough for
    a demo. I did add more comments than usual to make it easier to walk through the code. Scroll to the bottom of this
    script to see where it all gets started.

**WARNING**

    This module was minimally tested for https type certificates only.

**Data Structures**

    A list of the dictionaries below is created in _get_certs(). These dictionaries are the principal data structure
    used by all methods importing this module and the cert_update.py module.

    +-----------------------+-------+-------------------------------------------------------------------------------+
    | Key                   | Type  | Value                                                                         |
    +=======================+=======+===============================================================================+
    | certificate-entity    | str   | Same as returned from 'certificate-entity'                                    |
    +-----------------------+-------+-------------------------------------------------------------------------------+
    | certificate-type      | str   | Same as returned from 'certificate-type'                                      |
    +-----------------------+-------+-------------------------------------------------------------------------------+
    | certificate           | str   | 'certificate', if present, is a string containing several parameters. The     |
    |                       |       | format defined by a crypto standards body. The details of which are beyond    |
    |                       |       | the scope of this script. It is simply copied from the 'certificate' object   |
    |                       |       | returned from the API.                                                        |
    +-----------------------+-------+-------------------------------------------------------------------------------+
    | certificate-hexdump   | str   | The same comments as with certificate apply to certificate-hexdump. The       |
    |                       |       | certificate expiration dates are extracted from it and used in cert_control.  |
    |                       |       | Since this is a standard format, there is a Python library to decrypt it.     |
    |                       |       | Typical code use is:                                                          |
    |                       |       | from cryptography import x509                                                 |
    |                       |       | x509.load_pem_x509_certificate(cert_d['certificate-hexdump'])                 |
    +-----------------------+-------+-------------------------------------------------------------------------------+
    | certificate-verbose   | str   | The same comments as with 'certificate' apply                                 |
    +-----------------------+-------+-------------------------------------------------------------------------------+
    | cert_control          | dict  | Added in get_certs(). See cert_control below for detail.                      |
    +-----------------------+-------+-------------------------------------------------------------------------------+

    cert_control

    +---------------+-------+---------------------------------------------------------------------------+
    | Key           | Type  | Value                                                                     |
    +===============+=======+===========================================================================+
    | update        | bool  | If True, the cert exists and meets the criteria for update.               |
    +---------------+-------+---------------------------------------------------------------------------+
    | missing       | bool  | If True, the cert specified to be checked for an update does not exists   |
    |               |       | on the switch yet.                                                        |
    +---------------+-------+---------------------------------------------------------------------------+
    | begins        | str   | Human readable time stamp for when the cert is valid                      |
    +---------------+-------+---------------------------------------------------------------------------+
    | begins_epoch  | float | Same as begins in epoch time.                                             |
    +---------------+-------+---------------------------------------------------------------------------+
    | expires       | str   | Human readable time stamp for when the cert expires                       |
    +---------------+-------+---------------------------------------------------------------------------+
    | expires_epoch | float | Same as expires in epoch time.                                            |
    +---------------+-------+---------------------------------------------------------------------------+
    | new_cert      | str   | Updated in certs_add.py. Contains just the certificate portion of the     |
    |               |       | PEM file.                                                                 |
    +---------------+-------+---------------------------------------------------------------------------+

**Requirements**

    * FOS 9.1.0 or higher
    * Python 3.3 or higher

Version Control::

    +-----------+---------------+-----------------------------------------------------------------------------------+
    | Version   | Last Edit     | Description                                                                       |
    +===========+===============+===================================================================================+
    | 1.0.0     | 28 Apr 2022   | Initial launch                                                                    |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 1.0.1     | 26 Mar 2022   | Fixed FOS requirement in description.                                             |
    +-----------+---------------+-----------------------------------------------------------------------------------+
    | 1.0.2     | 09 May 2023   | used brcdapi_rest.operations_request() in decommission_port()                     |
    +-----------+---------------+-----------------------------------------------------------------------------------+
"""

__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2022, 2023 Jack Consoli'
__date__ = '09 May 2023'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack.consoli@broadcom.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '1.0.2'

import datetime
import argparse
from cryptography import x509
from cryptography.hazmat.backends import default_backend
import urllib3
import openpyxl as xl
import openpyxl.utils.cell as xl_utils
import collections
import time
import brcdapi.log as brcdapi_log
import brcdapi.excel_fonts as excel_fonts
import brcdapi.brcdapi_rest as brcdapi_rest
import brcdapi.fos_auth as fos_auth
import brcdapi.file as brcdapi_file
import brcdapi.excel_util as excel_util
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util

_DOC_STRING = False  # Should always be False. Prohibits any code execution. Only useful for building documentation
_DEBUG = False  # Intended for use with development tools. When True, use _DEBUG_xxx below instead of passed arguments
_DEBUG_i = 'test/test_input'
_DEBUG_a = 'eval'
_DEBUG_r = None
_DEBUG_d = False
_DEBUG_sup = False
_DEBUG_log = '_logs'
_DEBUG_nl = False

_SECONDS_PER_DAY = 60 * 60 * 24  # 60 sec/min, 60 min/hour, 24 hours/day
_cert_keys = ('certificate-entity', 'certificate-type')
_expiration_keys = ('begins', 'expires')  # in cert_control
_report_names = ('full', 'missing', 'update')  # Used in _create_report()
_REPORT_TYPE_FULL = 0
_REPORT_TYPE_MISSING = 1
_REPORT_TYPE_UPDATE = 2
_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_align_wrap = excel_fonts.align_type('wrap')
_align_wrap_c = excel_fonts.align_type('wrap_center')
_border_thin = excel_fonts.border_type('thin')
_used_files = dict()  # Used to make sure we don't overwrite a CSR or cert file
_begin_obj = '-----BEGIN'
_end_obj = '-----END'
_end = '-----'
_alt_name = {'subject-alternative-name-dns-names': 'dns-name', 'subject-alternative-name-ip-addresses': 'ip-address'}
"""The key in _report_defaults below is the report column header. The associated dictionary is defined as follows:
+-------+-------+---------------------------------------------------------------------------------------------------+
| key   | type  | Description                                                                                       |
+=======+=======+===================================================================================================+
| c     | int   | Default column width. Used when a column width was not returned from the openpyxl library         |
+-------+-------+---------------------------------------------------------------------------------------------------+
| v     | str,  | Default value to put in the report if none was present in the input file. PUT YOUR OWN DEFAULT IN |
|       | None  | HERE                                                                                              |
+-------+-------+---------------------------------------------------------------------------------------------------+
| str   | bool  | If True, convert the value to text                                                                |
+-------+-------+---------------------------------------------------------------------------------------------------+
| l     | bool  | If True, this is a login credential.                                                              |
+-------+-------+---------------------------------------------------------------------------------------------------+
"""
_report_defaults = collections.OrderedDict()
_report_defaults['ip_addr'] = dict(c=14, v=None, str=False, l=True)
_report_defaults['user_id'] = dict(c=14, v=None, str=False, l=True)
_report_defaults['pw'] = dict(c=14, v=None, str=False, l=True)
_report_defaults['security'] = dict(c=10, v=None, str=False, l=True)
_report_defaults['certificate-entity'] = dict(c=12, v=None, str=False, l=False)
_report_defaults['certificate-type'] = dict(c=12, v=None, str=False, l=False)
_report_defaults['days'] = dict(c=8, v=None, str=False, l=False)
_report_defaults['begins'] = dict(c=12, v=None, str=False, l=False)
_report_defaults['expires'] = dict(c=12, v=None, str=False, l=False)
_report_defaults['in_file'] = dict(c=18, v=None, str=False, l=False)
_report_defaults['out_file'] = dict(c=18, v=None, str=False, l=False)
_report_defaults['algorithm-type'] = dict(c=14, v='rsa', str=False, l=False)
_report_defaults['key-size'] = dict(c=10, v='2048', str=True, l=False)
_report_defaults['hash-type'] = dict(c=8, v='sha256', str=False, l=False)
_report_defaults['years'] = dict(c=8, v=1, str=False, l=False)
_report_defaults['country-name'] = dict(c=8, v='US', str=False, l=False)  # Must be 2 characters
_report_defaults['state-name'] = dict(c=14, v='FL', str=False, l=False)
_report_defaults['locality-name'] = dict(c=14, v='St. Augustine', str=False, l=False)
_report_defaults['organization-name'] = dict(c=14, v='Pre-sales SE', str=False, l=False)
_report_defaults['unit-name'] = dict(c=14, v='BSN', str=False, l=False)
_report_defaults['domain-name'] = dict(c=20, v='switch69.test2.com', str=False, l=False)
_report_defaults['subject-alternative-name-dns-names'] = dict(c=24, v=None, str=False, l=False)
_report_defaults['subject-alternative-name-ip-addresses'] = dict(c=24, v=None, str=False, l=False)
_login_keys = [k for k, d in _report_defaults.items() if d['l']]
_alt_names = ['subject-alternative-name-dns-names', 'subject-alternative-name-ip-addresses']
_param_keys = [k for k in _report_defaults.keys() if k not in _login_keys+_alt_names]
_report_hdr = {  # Key is the first row header. Dictionary is 's' for first column, 'e' last column for cell merge
    'Login Credentials': dict(s='ip_addr', e='security'),
    'Input Parameters': dict(s='certificate-entity', e='days'),
    'Certificate Expiry': dict(s='begins', e='expires'),
    'Certificate Files': dict(s='in_file', e='out_file'),
    'Required for CSR': dict(s='algorithm-type', e='domain-name'),
    'Optional CSR Input': dict(s='subject-alternative-name-dns-names', e='subject-alternative-name-ip-addresses'),
}


def _extract_certificate(file):
    """Extracts the certificates and CSRs from a PEM file.

    Depending on how well controlled your environment is, this may not be needed.  What needs to be sent to the switch
    is just the certificate or CSR portion but the CA administrator may have given you a file containing verbose output.
    If that happened, the plain text information must be stripped out. The only thing FOS wants to see is the base64
    encoded certificate + the beginning and ending demarcations. Specifically, this is everything between '-----BEGIN'
    and '-----END xxx-----' inclusive of '-----BEGIN', the trailing '-----' in '-----END xxx-----'.

    :param file: Standard PEM file.
    :type file: str
    :return: The certificates parsed from the file contents. None if an error occured or no certificated found
    :rtype: str, None
    """
    global _begin_obj, _end_obj, _end

    # Read the file
    try:
        f = open(file, 'rb')
        # FOS only accepts Unix style new lines
        buf = f.read().decode(brcdapi_util.encoding_type, errors='ignore').replace('\r', '')
        f.close()
    except FileNotFoundError:
        brcdapi_log.log('  File not found: ' + file, True)
        return None

    # Parse out just the certificates & keys
    r_buf, begin_i = '', buf.find(_begin_obj)
    while begin_i >= 0:
        end_i = buf.find(_end_obj)
        if end_i < begin_i:
            brcdapi_log.log('Corrupted PEM file. Mismatched ' + _begin_obj + ' and ' + _end_obj + ' in ' + file, True)
            return None
        end_i += len(_end_obj)
        end_i += buf[end_i:].find(_end) + len(_end)
        r_buf += buf[begin_i: end_i] + '\n'  # I don't think FOS needs this '\n' but it makes it easier to read
        buf = buf[end_i:]
        begin_i = buf.find(_begin_obj)

    return r_buf if len(r_buf) > 0 else None


def _matching_param(param_l, cert_d):
    """A utilitarian method to find a parameter that matches a certificate returned in the API

    :param param_l: List of parameters ('param_l' in input_d returned from _get_input())
    :type param_l: list
    :param cert_d: The certificate as returned from the API
    :type cert_d: dict
    :return: The dictionary entry in param_l that matches the certificate in cert_d. None if not found.
    :rtype: dict, None
    """
    global _cert_keys

    for param_d in param_l:
        match_count = 0
        for key in _cert_keys:
            if cert_d.get(key) is not None and param_d.get(key) is not None and cert_d.get(key) == param_d.get(key):
                match_count += 1
                if match_count == len(_cert_keys):
                    return param_d

    return None


def _matching_cert(cert_l, param_d):
    """A utilitarian method to find a parameter that matches a certificate returned in the API

    :param cert_l: List of certificates ('certs_l' in input_d returned from _get_certs() as returned from the API)
    :type cert_l: list
    :param param_d: The parameter as returned from the _get_inputs in input_d['param_l']
    :type param_d: dict
    :return: The dictionary entry in certs_l that matches the certificate in cert_d. None if not found.
    :rtype: dict, None
    """
    global _cert_keys

    for cert_d in cert_l:
        if cert_d['certificate-entity'] == param_d['certificate-entity'] and \
                cert_d['certificate-type'] == param_d['certificate-type']:
            return cert_d

    return None


def _certs_filter(switch_d, filter_type):
    """Filters the list of certs based on the filter type

    :param switch_d: Switch dictionary from input_d[ip_addr] as returned from _get_input()
    :type switch_d: dict
    :param filter_type: 0 - All certs, 1 - Missing certs, 2 - certs due for an update
    :type filter_type: int
    :return: Filtered list of certs
    :rtype: list
    """
    if filter_type == 0:
        return switch_d['_certs_l']
    elif filter_type == 1:
        return [cert_d for cert_d in switch_d['_certs_l'] if cert_d['cert_control']['missing']]
    elif filter_type == 2:
        return [cert_d for cert_d in switch_d['_certs_l'] if cert_d['cert_control']['update']]
    brcdapi_log.exception('Invalid filter_type: ' + str(filter_type), True)
    return list()


def _create_report(input_d, report_type):
    """Creates a report in an Excel workbook

    :param input_d: As returned from _get_input()
    :type input_d: dict
    :param report_type: Index into _report_names
    :type report_type: int
    """
    global _report_names, _login_keys, _cert_keys, _expiration_keys, _std_font, _bold_font, _align_wrap, _align_wrap_c
    global _border_thin, _used_files, _report_hdr, _report_defaults, _REPORT_TYPE_FULL, _alt_names

    # Create the workbook
    report = input_d.get('_report')
    if report is None:
        return
    report_name = report[0: len(report)-len('.xlsx')] + '_' + _report_names[report_type] + '.xlsx'
    wb = xl.Workbook()
    sheet = wb.create_sheet(index=0, title='parameters')

    # Add the second row header and columns to _report_defaults
    row, col = 2, 1
    for k, d in _report_defaults.items():
        d.update(col=col)
        sheet.column_dimensions[xl_utils.get_column_letter(col)].width = d['c']
        excel_util.cell_update(sheet, row, col, k, font=_bold_font, align=_align_wrap_c, border=_border_thin)
        col += 1

    # Add the first header and set the cell merge
    row, col = 1, 1
    for k, d in _report_hdr.items():
        col = _report_defaults[d['s']]['col']
        sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=_report_defaults[d['e']]['col'])
        excel_util.cell_update(sheet, row, col, k, font=_bold_font, align=_align_wrap, border=_border_thin)

    # The intent of switch_l is to contain a filtered list of switch dictionaries that belong in the report. report_l is
    # a list of dictionaries matching the content of switch_l in the format used to display the certificates in the
    # workbook.
    report_l, switch_l = list(), list()

    # Determine what to put in the report. This is the filtered list of switch dictionaries put in switch_l
    for switch_d in [d for k, d in input_d.items() if k[0] != '_']:
        if report_type == _REPORT_TYPE_FULL:
            switch_l.append(switch_d)
        else:
            # The value at the index (which is report_type) into _report_names is also the name of the flag in
            # cert_control that is used to determine if a cert is missing or needs to be updated.
            filtered_switch_l = [c for c in switch_d['_certs_l'] if c['cert_control'][_report_names[report_type]]]
            if len(filtered_switch_l) > 0:
                filtered_switch_d = dict(_certs_l=filtered_switch_l)
                for k, v in switch_d.items():
                    if k != '_certs_l':
                        filtered_switch_d.update({k: v})
                switch_l.append(filtered_switch_d)

    # Format the content of switch_l into dictionaries formatted for the workbook. These are added to report_l
    for switch_d in switch_l:
        report_d = dict()
        for key in _login_keys:  # Login credentials are only added to the report once for each switch.
            report_d.update({key: switch_d[key]})
        for cert_d in switch_d['_certs_l']:  # Now add all the certificates
            for key in _cert_keys:
                buf = _report_defaults[key]['v'] if cert_d.get(key) is None else cert_d[key]
                report_d.update({key: buf})
            for key in _expiration_keys:
                buf = cert_d['cert_control'].get(key)
                if buf is None:
                    buf = _report_defaults[key]['v']  # Future proofing. _report_defaults[key]['v'] is always None
                report_d.update({key: buf})
            param_d = _matching_param(switch_d['_params_l'], cert_d)
            if param_d is None:
                param_d = dict()
            for p_key in _param_keys + _alt_names:
                if p_key not in _cert_keys + _expiration_keys:
                    buf = param_d.get(p_key) if report_type == _REPORT_TYPE_FULL else \
                        _report_defaults[p_key]['v'] if param_d.get(p_key) is None else _report_defaults[p_key]['v']
                    report_d.update(({p_key: buf}))
            report_l.append(report_d.copy())
            report_d = dict()

    # Add each switch and certificate to the report
    row = 3
    for report_d in report_l:
        for key in _report_defaults.keys():
            v = report_d.get(key)
            buf = ';'.join(v) if isinstance(v, list) else v
            excel_util.cell_update(sheet, row, _report_defaults[key]['col'], buf, font=_std_font, align=_align_wrap,
                                   border=_border_thin)
        row += 1

    # Write out the report
    try:
        wb.save(report_name)
    except FileNotFoundError:
        brcdapi_log.log('Folder in path ' + report_name + ' does not exist.', True)
    except PermissionError:
        brcdapi_log.log('Could not write ' + report_name + ' because it is open in another application.', True)


def _eval_certs(switch_d):
    """Checks for missing and about to expire certificates

    :param switch_d: Switch dictionary from input_d[ip_addr] as returned from _get_input()
    :type switch_d: dict
    """
    global _SECONDS_PER_DAY

    for param_d in switch_d['_params_l']:  # For each certificate defined in the input workbook
        cert_d = _matching_cert(switch_d['_certs_l'], param_d)  # Find the matching cert
        if cert_d is None:
            # I'm not validating what was put in the workbook. Keep in mind, using the workbook was an expedient for
            # testing purposes. The normal reason for not finding a match is when the action is "eval" but no parameters
            # were defined. The content of everything in param_d will be None so no match will be found. Running this
            # script with no parameters as input to eval generates a report with all the current certificate information
            # but there is nothing to evaluate so "missing" and "update" reports will have nothing in them.
            continue

        # Find the matching cert in the list of certs returned from the switch
        if isinstance(param_d.get('days'), int):
            expire = cert_d['cert_control'].get('expires_epoch')
            if isinstance(expire, float):
                if expire - datetime.datetime.now().timestamp() - param_d['days']*_SECONDS_PER_DAY <= 0:
                    cert_d['cert_control']['update'] = True
            else:
                cert_d['cert_control']['missing'] = True
        else:
            buf = '  Could not evaluate because "days" was not specified in the input workbook for '
            buf += cert_d['certificate-entity'] + ', ' + cert_d['certificate-type']
            brcdapi_log.log(buf, True)

    return


def _del_cert(session, entity, cert_type):
    """Deletes a certificate

    :param session: Session object returned from brcdapi.fos_auth.login()
    :type session: dict
    :param entity: Certificate entity
    :type entity: str
    :param cert_type: Certificate type
    :type cert_type: str
    :return: True: Successfully deleted the certificate. False: An error occured while attempting to delete the cert.
    :rtype: bool
    """
    content = {
        'security-certificate-action': {
            'certificate-entity': entity,
            'certificate-type': cert_type
        }
    }
    obj = brcdapi_rest.send_request(session,
                                    'running/brocade-security/security-certificate-action',
                                    'DELETE',
                                    content)
    if fos_auth.is_error(obj):
        brcdapi_log.log(['  Error deleting '+entity+', '+cert_type, '    '+fos_auth.formatted_error_msg(obj)], True)
        return False

    return True


def _generate_csr(session, param_d):
    """Generate a certificate signing request (CSR)

    :param session: Session object returned from brcdapi.fos_auth.login()
    :type session: dict
    :param param_d: One of the dictionaries in the list of parameters ('param_l' in input_d returned from _get_input())
    :type param_d: dict
    :return: True: Successfully generated the CSR. False: An error occured while attempting to generate the CSR.
    :rtype: bool
    """
    global _alt_name

    sub_content = {  # Put your own defaults in here
        'certificate-entity': 'csr',
        'certificate-type': param_d['certificate-type'],
        'algorithm-type': 'rsa' if param_d.get('algorithm-type') is None else param_d.get('algorithm-type'),
        'key-size': '2048' if param_d.get('key-size') is None else str(param_d.get('key-size')),
        'hash-type': 'sha256' if param_d.get('hash-type') is None else param_d.get('hash-type'),
        'years': 1 if param_d.get('years') is None else param_d.get('years'),
        'country-name': 'US' if param_d.get('country-name') is None else param_d.get('country-name'),
        'state-name': 'CA' if param_d.get('state-name') is None else param_d.get('state-name'),
        'locality-name': 'San Jose' if param_d.get('locality-name') is None else param_d.get('locality-name'),
        'organization-name': 'Pre Sales' if param_d.get('organization-name') is None else \
            param_d.get('organization-name'),
        'unit-name': 'BSN' if param_d.get('unit-name') is None else param_d.get('unit-name'),
        'domain-name': 'brm.bsnlab.broadcom.net' if param_d.get('domain-name') is None else param_d.get('domain-name')
    }
    # Add optional subject alternatives
    for key, sub_key in _alt_name.items():
        alt_subj = param_d.get(key)
        if len(alt_subj) > 0:
            sub_content.update({key: {sub_key: alt_subj}})
    content = {'security-certificate-generate': sub_content}

    # Send the CSR to the switch
    obj = brcdapi_rest.send_request(session,
                                    'running/brocade-security/security-certificate-generate',
                                    'POST',
                                    content)
    if fos_auth.is_error(obj):
        brcdapi_log.log(['  Error generating CSR for ' + param_d['certificate-entity']+', '+param_d['certificate-type'],
                         '    ' + fos_auth.formatted_error_msg(obj)], True)
        return False

    return True


def _extract_cert_or_csr(switch_d, entity, cert_type, out_file):
    """Extracts a certificate or CSR and writes it to a file

    :param switch_d: Switch dictionary from input_d[ip_addr] as returned from _get_input()
    :type switch_d: dict
    :param entity: Certificate entity
    :type entity: str
    :param cert_type: Certificate type
    :type cert_type: str
    :param out_file: Name of output file. May include pathing.
    :type out_file: str
    :return: True: Successfully exported the certificate. False: An error occured while attempting to export the cert.
    :rtype: bool
    """
    new_cert_l = _get_certs(switch_d)
    for n_cert_d in gen_util.convert_to_list(new_cert_l):  # If there was an error reading it, _get_certs() reported it.
        if n_cert_d['certificate-entity'] == entity and n_cert_d['certificate-type'] == cert_type:
            hexdump = n_cert_d.get('certificate-hexdump')
            if isinstance(hexdump, str) and len(hexdump) > 0:
                brcdapi_log.log('  Writing ' + entity + ', ' + cert_type + ' to ' + out_file, True)
                try:
                    with open(out_file, 'w') as f:
                        f.write(hexdump)
                        return True
                except FileNotFoundError:
                    brcdapi_log.log('  Folder in ' + out_file + ' not found.', True)
                    return False
            else:
                brcdapi_log.log('  No certificate associated with ' + entity + ', ' + cert_type, True)
                return False

    brcdapi_log.log('  ' + entity + ', ' + cert_type + ' does not exist.', True)
    return False


def _add_cert(session, entity, cert_type, cert_file):
    """Extracts a certificate or CSR and writes it to a file

    :param session: Session object returned from brcdapi.fos_auth.login()
    :type session: dict
    :param entity: Certificate entity
    :type entity: str
    :param cert_type: Certificate type
    :type cert_type: str
    :param cert_file: Name of file containing the certificate generated from a CA.
    :type cert_file: str
    :return: True: Successfully added the certificate. False: An error occured while attempting to add the cert.
    :rtype: bool
    """
    hexdump = _extract_certificate(cert_file)
    if hexdump is None:
        brcdapi_log.log('  No certificates found in ' + cert_file, True)
        return False

    content = {
        'security-certificate-parameters': {
            'certificate-entity': entity,
            'certificate-type': cert_type,
            'action': 'import',
            'certificate-hexdump': hexdump
        }
    }
    obj = brcdapi_rest.operations_request(session, 'operations/security-certificate', 'POST', content)
    if fos_auth.is_error(obj):
        brcdapi_log.log(['  Error adding ' + entity + ', ' + cert_type, '    '+fos_auth.formatted_error_msg(obj)], True)
        return False

    return True


###################################################################
#
#             Action Methods used in _valid_actions
#
###################################################################
def _eval_action(switch_d):
    """Evaluates certificates

    :param switch_d: Switch dictionary from input_d[ip_addr] as returned from _get_input()
    :type switch_d: dict
    """
    global _report_names

    _eval_certs(switch_d)
    for i in range(0, len(_report_names)):
        _create_report(switch_d['_parent'], i)


def _add_cert_action(switch_d):
    """Applies the certificates. See _eval_action() for parameter details"""
    for param_d in switch_d['_params_l']:
        cert_file = param_d.get('in_file')
        if isinstance(cert_file, str) and len(cert_file) > 0:
            csr_file = brcdapi_file.full_file_name(cert_file, '.pem')
            cert_entity, cert_type = param_d['certificate-entity'], param_d['certificate-type']
            brcdapi_log.log('  Adding ' + csr_file + ' for ' + cert_entity + ', ' + cert_type, True)
            if _add_cert(switch_d['_session'], cert_entity, cert_type, csr_file):
                brcdapi_log.log('  Successfully added ' + csr_file + ' for ' + cert_entity + ', ' + cert_type, True)
        else:
            brcdapi_log.log('  Missing cert file for ' + param_d['certificate-entity'] + ', ' +
                            param_d['certificate-type'], True)


def _del_cert_action(switch_d):
    """Deletes the certificates. See _eval_action() for parameter details"""
    session = switch_d['_session']
    param_l = switch_d['_params_l']

    # If other certs are present, delete the cert (private key) first, then the public keys, then the CSR.
    ordered_l = [param_d for param_d in param_l if param_d['certificate-entity'] == 'cert']
    ordered_l.extend([param_d for param_d in param_l if param_d['certificate-entity'] in ('ca-server', 'ca-client')])
    ordered_l.extend([param_d for param_d in param_l if param_d['certificate-entity'] == 'csr'])
    for param_d in ordered_l:
        cert_d = _matching_cert(switch_d['_certs_l'], param_d)
        if cert_d is None:
            continue  # This happens when there are alternative subject names because the entity & type are None
        param_entity = param_d['certificate-entity']
        param_type = param_d['certificate-type']
        hexdump = cert_d.get('certificate-hexdump')
        if isinstance(hexdump, str) and len(hexdump) > 0:

            # Delete the certificate
            brcdapi_log.log('  Deleting ' + param_entity + ', ' + param_type, True)
            if not _del_cert(session, param_entity, param_type):
                raise IOError  # An error message is logged in _del_cert() if this occurs so just bail out.

            # If the HTTPs cert was deleted, the interface reverts back to HTTP so we will need to re-login
            if param_entity == 'cert' and param_type == 'https':
                # It typically takes 5-6 seconds for the interface to be restored after deleting a certificate. The
                # driver will retry if the interface is busy so the sleep below is not necessary but rather than
                # immediately hammering the switch with additional requests, I thought giving it a chance to recover
                # seemed like the right thing to do.
                time.sleep(6)
                switch_d['security'] = 'none'
                _login(switch_d)
                if switch_d['_session'] is None:
                    raise IOError  # An error message is logged in _login() if this occurs so just bail out.
        else:
            brcdapi_log.log('  Cert does not exist: ' + param_entity + ', ' + param_type, True)
            continue


def _csr_action(switch_d):
    """Generates a CSR for each certificate in switch_d. See _eval_action() for parameter details"""
    global _used_files

    session = switch_d['_session']
    for param_d in switch_d['_params_l']:

        """Programmer Notes:
        
        Step 1: Make sure you have some place to export the CSR to before generating the CSR. Your organization may want
        you to create a name for the CSR file based on naming conventions but in this example names are manually entered
        into the workbook read in by this script. To help minimize effort for the demonstration this script was used
        for, _eval_action() automatically pre-populates the workbook with file names. More thorough error checking would
        make sure that any folders specified in the name exist before continuing but in this exercise, that
        determination is made with a simple try/except after the CSR was generated. Your organization may have a means
        to automate generating a CSR that may not require you to export the CSR to a file. 
        
        Step 2: Before generating a CSR, it's a good practice to delete existing certificates before generating a CSR.
        Furthermore, certificates should be deleted in the order they were created.
        
        Step 3: Generate the CSR.
        
        Step 4: Read the CSR and write it out to the file created in Step 1. Note that the CSR is not returned with the
        Rest request to generate the CSR. It must be read back. If multiple CSRs are being generated on the same switch,
        it would be more efficient to generate all the CSRs, then read them all back but this example takes a simple one
        at a time approach."""

        # Step 1: Make sure there is a place to put the CSR before taking any action.
        csr_file = param_d.get('out_file')
        if csr_file is None:
            e_buf = 'WARNING: No output file for ' + switch_d['ip_addr'] + ', ' + param_d['certificate-type']
            e_buf += '. a CSR was not generated.'
            brcdapi_log.log(['', e_buf, ''], True)
            return
        csr_file = brcdapi_file.full_file_name(csr_file, '.csr')

        # Step 2: Before generating a CSR, check for existing certificates and if they exist, delete them. While
        # reviewing the code and adding comments I realized that I'm only deleting the cert. Since a CSR is being
        # generated, the associated ca-server should be deleted. Deleting the ca-server cert is best practice but it's
        # going to get over written. I figured this was good enough for what this module was intended for. I already
        # finished testing and didn't want to circle back to more testing.
        brcdapi_log.log('  Checking for exising certs to remove before generating the CSR', True)
        _del_cert_action(switch_d)

        # Step 3: Generate the CSR
        brcdapi_log.log('  Generating ' + param_d['certificate-entity'] + ', ' + param_d['certificate-type'], True)
        if not _generate_csr(session, param_d):
            return  # Not sure what is wrong with this switch so return without causing any more damage

        # Step 4: Read the CSR back and write it out to a file
        _extract_cert_or_csr(switch_d, 'csr', param_d['certificate-type'], csr_file)


""" Key in _valid_actions is the -a input parameter, returned as _action in dict returned from _get_input(). In the
sub-dict, 'd' is the description used in _get_input() to generate a help message. 'a' is a pointer to the method to
called in pseudo_main().

Programmer's tip: Python has many attributes of more advanced programming languages such as C++. For many tools team
programmers used to more basic scripting languages, calling a method from a table may be unusual but this is very common
for table driven software. For this table, the pointer to the method is loaded by omitting the parenthesis."""
_valid_actions = dict(
    eval=dict(d='Reads certificates and generate reports.', a=_eval_action),
    csr=dict(d='Generates a certificate signing request (CSR). Deletes certificate if present', a=_csr_action),
    add_cert=dict(d='Adds certificates.', a=_add_cert_action),
    del_cert=dict(d='Deletes certificates if present.', a=_del_cert_action),
)


def _get_certs(switch_d):
    """Reads the certificates from the API and adds control_d, see cert_control in module header.

    :param switch_d: Switch dictionary from input_d[ip_addr] as returned from _get_input()
    :type switch_d: dict
    :return: List of cert dictionaries. See **Data Structures** in the module header.
    :rtype: list
    """
    session = switch_d['_session']
    brcdapi_log.log('  Reading certificates. This will take 40-60 sec.', True)

    # Get the certificates from the API
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)  # Disable self-signed cert warnings
    obj = brcdapi_rest.get_request(session, 'running/brocade-security/security-certificate')
    if fos_auth.is_error(obj):
        brcdapi_log.exception(['  Failed to capture certificates.', '    ' + fos_auth.formatted_error_msg(obj)], True)
        return None

    # Parse the certificates we got back from the API, check for updates, and add to cert_control (control_d)
    cert_l = list()
    for cert_d in gen_util.convert_to_list(obj.get('security-certificate')):  # It should always be a list
        cert_l.append(cert_d)
        # For details on control_d, see cert_control in module header
        control_d = dict(update=False, missing=False, new_cert=None)
        cert_d.update(cert_control=control_d)

        # Get the certificate dates
        # hexdump below should always be a str. I'm extra cautious with data from an external source. Keep in mind that
        # this content is from a library and not derived by FOS itself.
        hexdump = cert_d.get('certificate-hexdump')
        if not isinstance(hexdump, str):  # Do a cert_d.get below because I have no idea what went wrong if we get here
            buf = '  Invalid certificate for ' + str(cert_d.get('certificate-entity')) + ', '
            buf += str(cert_d.get('certificate-type'))
            brcdapi_log.log(buf, True)
            hexdump = ''
        if len(hexdump) > 0:
            # With the cryptography library 3.1 and above, default_backend() is used if not specified. Earlier versions
            # require it explicitly so below is intended to work with both.
            try:
                cert = x509.load_pem_x509_certificate(hexdump.encode(), default_backend())
                not_valid_before = cert.not_valid_before
                not_valid_after = cert.not_valid_after

                # Below uses the current time of the server running the script. Good enough for time measured in days
                control_d.update(begins=not_valid_before.strftime('%d %b %Y'),
                                 begins_epoch=not_valid_before.timestamp(),
                                 expires=not_valid_after.strftime('%d %b %Y'),
                                 expires_epoch=not_valid_after.timestamp())

            except ValueError:
                pass  # There isn't a certificate. I think the only time we get here is if it's just a CSR

    return cert_l


def _login(switch_d):
    """Logs into the switch

    :param switch_d: Switch dictionary from input_d[ip_addr] as returned from _get_input()
    :type switch_d: dict
    """
    sec = 'none' if switch_d['security'] is None else switch_d['security']
    brcdapi_log.log('  Attempting login', True)
    session = brcdapi_rest.login(switch_d['user_id'], switch_d['pw'], switch_d['ip_addr'], sec)
    if fos_auth.is_error(session):
        brcdapi_log.log(['  Login failed. Error message is:', '    ' + fos_auth.formatted_error_msg(session)], True)
        switch_d['_session'] = None
    else:
        brcdapi_log.log('  Login succeeded', True)
        switch_d['_session'] = session


def _logout(switch_d):
    """Logs into the switch

    :param switch_d: Switch dictionary from input_d[ip_addr] as returned from _get_input()
    :type switch_d: dict
    """
    session = switch_d.get('_session')
    if session is not None:
        brcdapi_log.log('  Attempting logout', True)
        obj = brcdapi_rest.logout(session)
        if fos_auth.is_error(obj):
            brcdapi_log.log(['  Logout failed. Error message is:', '    ' + fos_auth.formatted_error_msg(obj)], True)
        else:
            brcdapi_log.log('  Logout succeeded.', True)


def _get_input():
    """Parses the module load command line, performs basic parameter validation checks, and sets up the log file.

    Programmer's tip: The intent of this module is to illustrate how to manage security certificates. Doing that
    requires more input than is reasonable to put in a command string so I read data from a Workbook reusing utility
    methods I wrote for other applications. Since the nature of how certificates are generated varies widely for
    different organizations, I didn't spend much time on making this neat and more importantly, I didn't spend much time
    validating the user input.

    The return dictionary has a key/value pair as follows:
    
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | Key       | Type  | Description                                                                               |
    +===========+=======+===========================================================================================+
    | _report   | str   | The -r option from the command shell. This is the report prefix.                          |
    |           | None  |                                                                                           |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | _action   | str   | The -a option from the command shell. An error is returned if this parameters is not in   |
    |           |       | the actions defined in _valid_actions                                                     |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | ip_addr*  | dict  | A dictionary of switch dictionaries. The keys are the switch IP address. The values are   |
    |           |       | as defined in switch_d below.                                                             |
    +-----------+-------+-------------------------------------------------------------------------------------------+

    switch_d
    
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | Key       | Type  | Description                                                                               |
    +===========+=======+===========================================================================================+
    | user_id   | str   | Login ID                                                                                  |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | pw        | str   | Password                                                                                  |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | security  | str   | Security, 'none', 'self', or the CA                                                       |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | _params_l | list  | List of dictionaries for all remaining row items. This is a list because subsequent rows  |
    |           |       | with no login credentials are more certs associated with the same switch.                 |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | _session  | dict  | Session object as returned from brcdapi.fos_auth.login()                                  |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | _parent   | dict  | Pointer to the parent object. Returned as rd from this method.                            |
    +-----------+-------+-------------------------------------------------------------------------------------------+

    :return ec: Exit code. 0: OK. -1: Errors found
    :rtype ec: int
    :return rd: See description above in the method description
    :rtype rd: dict
    """
    global _DEBUG_i, _DEBUG_a, _DEBUG_r, _DEBUG_sup, _DEBUG_d, _DEBUG_log, _DEBUG_nl, _login_keys, _valid_actions
    global _report_defaults, _param_keys, _alt_names

    ec, rd = 0, dict()

    if _DEBUG:
        args_i, args_a, args_r, args_sup, args_d, args_log, args_nl = \
            _DEBUG_i, _DEBUG_a, _DEBUG_r, _DEBUG_sup, _DEBUG_d, _DEBUG_log, _DEBUG_nl

    else:
        buf = 'Typically used with action eval first which checks for expiring certificates and generates a workbook '\
              'that can be modified and used as input for other actions.'
        parser = argparse.ArgumentParser(description=buf)
        buf = '(Required) Excel file of login credentials and certs to check for. See certs_get_example.xlsx. '\
              '".xlsx" is automatically appended if no extension is specified.'
        parser.add_argument('-i', help=buf, required=True)
        buf = '(Required) Action to take. Options are:'
        for k, d in _valid_actions.items():
            buf += ' ' + str(k) + ': ' + d['d']
        parser.add_argument('-a', help=buf, required=True)
        buf = '(Required if the action, -a, is "eval") Report name prefix. Three reports are generated. The reports '\
              'are named by appending this with "_full", "_missing", and "_update". May include a folder.'
        parser.add_argument('-r', help=buf, required=False)
        parser.add_argument('-sup', help=buf, action='store_true', required=False)
        buf = '(Optional) No parameters. When set, a pprint of all content sent and received to/from the API, except ' \
              'login information, is printed to the log.'
        parser.add_argument('-d', help=buf, action='store_true', required=False)
        buf = '(Optional) Directory where log file is to be created. Default is to use the current directory. The log '\
              'file name will always be "Log_xxxx" where xxxx is a time and date stamp.'
        parser.add_argument('-log', help=buf, required=False,)
        buf = '(Optional) No parameters. When set, a log file is not created. The default is to create a log file.'
        parser.add_argument('-nl', help=buf, action='store_true', required=False)
        args = parser.parse_args()
        args_i, args_a, args_r, args_sup, args_d, args_log, args_nl = \
            args.i, args.a, args.r, args.sup, args.d, args.log, args.nl

    # Set up the log file
    if args_sup:
        brcdapi_log.set_suppress_all()
    if not args_nl:
        brcdapi_log.open_log(args_log)
    if args_d:  # Verbose debug
        brcdapi_rest.verbose_debug = True

    # ml is a list of message to display with user feedback. Error messages may be added.
    ml = ['WARNING: Debug mode is enabled'] if _DEBUG else list()
    ml.append('Input file, -i:    ' + args_i)
    ml.append('Report prefix, -r: ' + str(args_r))
    ml.append('Action, -a:        ' + str(args_a))

    # Validate the input
    args_i = brcdapi_file.full_file_name(args_i, '.xlsx')
    if args_a not in _valid_actions:
        ml.append('Invalid action, -a: ' + args_a)
        ec = -1
    else:
        if args_r is None:
            if args_a == 'eval':
                ml.append('Action, -a, is "eval" but no report prefix, -r, was specified.')
                ec = -1
        else:
            rd.update(_report=brcdapi_file.full_file_name(args_r, '.xlsx'))
        rd.update(_action=args_a)
        # Read and parse the input file if specified
        try:
            input_d = excel_util.parse_parameters(wb_name=args_i, hdr_row=1)
        except FileNotFoundError:
            ml.append('Input file, -i, not found: ' + args_i)
            ec = -1

    if ec == 0:
        # Parse the input file - The desire is to login to each switch just once. Since human factors are involved in
        # editing the workbooks, the return dictionary is organized by switch.
        row, last_d, new_switch = 2, None, False
        for d in input_d['content']:

            # Get the login credentials if there are new login credentials, otherwise, use the previous credentials
            ip_d = last_d if d['ip_addr'] is None else rd.get(d['ip_addr'])
            if ip_d is None:
                if d['ip_addr'] is None:  # An initial row with an IP address was not found in the Workbook
                    ml.append('IP address unknown at row ' + str(row) + ' in ' + args_i)
                    ec = -1
                    continue
                ip_d = dict(_parent=rd, _params_l=list(), _certs_l=list(), _session=None)
                rd.update({d['ip_addr']: ip_d})
                for key in _login_keys:  # Add all the login keys
                    ip_d.update({key: d.get(key)})
                new_switch, last_d = True, ip_d

            # Add the parameters for this switch. I missed alternative subject names so shoe horned them in afterwards
            update_flag, temp_d = False, dict()
            for key in _param_keys:  # Add the parameters
                v = d.get(key)
                if v is not None:
                    update_flag = True
                temp_d.update({key: v})
            if update_flag or new_switch:
                last_param_d = param_d = temp_d.copy()
                for key in _alt_names:
                    param_d.update({key: list()})
                ip_d['_params_l'].append(param_d)
                new_switch = False
            for key in _alt_names:
                if d.get(key) is not None:
                    alt_l = last_param_d.get(key)
                    alt_l.extend(d[key].split(';'))

            row += 1

    # Update log with input parameters and echo back to STD_OUT
    if len(rd.keys()) == 0:
        ml.append('No content found in ' + args_i)
        ec = -1
    brcdapi_log.log(ml, True)

    return ec, rd


def pseudo_main():
    """Basically the main(). Did it this way so it can easily be modified to be called from another script.

    :return: Exit code. See exit codes in brcddb.brcddb_common
    :rtype: int
    """
    global _valid_actions

    # Get the command line input
    ec, input_d = _get_input()
    if ec != 0:
        return ec

    for ip_addr in [k for k in input_d.keys() if k[0] != '_']:
        brcdapi_log.log(['', 'Switch: ' + brcdapi_util.mask_ip_addr(ip_addr)], True)
        switch_d = input_d[ip_addr]

        # Login
        _login(switch_d)
        if switch_d['_session'] is None:
            ec = -1  # The error message is logged in _login()
            continue

        try:  # This try is to ensure the logout code gets executed regardless of what happens.
            switch_d['_certs_l'] = _get_certs(switch_d)  # Get and add the list of certs to switch_d
            _valid_actions[input_d['_action']]['a'](switch_d)
        except BaseException as e:
            brcdapi_log.exception('Programming error encountered. Exception is: ' + str(e), True)

        # Logout
        _logout(switch_d)  # Error messages, if any, are logged in _logout()

    return ec


###################################################################
#
#                    Main Entry Point
#
###################################################################
if _DOC_STRING:
    print('_DOC_STRING is True. No processing')
    exit(0)

_ec = pseudo_main()
brcdapi_log.close_log(['', 'All processing complete. Exit code: ' + str(_ec)], True)
exit(_ec)
