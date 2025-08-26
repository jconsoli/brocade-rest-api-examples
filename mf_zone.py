#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Copyright 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

**License**

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

**Description**

Creates a report in Excel Workbook format from a brcddb project

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 1.0.0     | 26 Dec 2024   | Initial Launch                                                                        |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 1.0.1     | 01 Mar 2025   | Help and error message enhancements.                                                  |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 1.0.2     | 25 Aug 2025   | Use brcddb.util.util.get_import_modules to dynamically determined imported libraries. |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2024, 2025 Consoli Solutions, LLC'
__date__ = '25 Aug 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '1.0.2'

import os
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.file as brcdapi_file
import brcdapi.excel_fonts as excel_fonts
import brcdapi.excel_util as excel_util
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util
import brcddb.brcddb_project as brcddb_project
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.brcddb_switch as brcddb_switch
import brcddb.brcddb_common as brcddb_common
import brcddb.brcddb_zone as brcddb_zone
import brcddb.util.iocp as brcddb_iocp
import brcddb.brcddb_port as brcddb_port
import brcddb.report.utils as report_utils

_DOC_STRING = False  # Should always be False. Prohibits any code execution. Only useful for building documentation
# _STAND_ALONE: True: Executes as a standalone module taking input from the command line. False: Does not automatically
# execute. This is useful when importing this module into another module that calls psuedo_main().
_STAND_ALONE = True  # See note above

_std_font = excel_fonts.font_type('std')
_bold_font = excel_fonts.font_type('bold')
_align_wrap = excel_fonts.align_type('wrap')
_zonecfg_col = (20, 68)


# For input parameter definitions, search for _input_d
class Found (Exception):
    pass


def _unique_alias_name(fabric_obj, alias_name):
    """Ensure that an alias name is unique by appending a number if alias_name is already used

    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :param alias_name: Alias name
    :type alias_name: str
    :return: Alias name
    :rtype: str
    """
    i, r_alias_name = 0, alias_name
    while fabric_obj.r_alias_obj(r_alias_name) is not None:
        r_alias_name = alias_name + '_' + str(i)

    return r_alias_name


def _unique_zone_name(fabric_obj, zone_name):
    """Ensure that a zone name is unique by appending a number if alias_name is already used

    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :param zone_name: Alias name
    :type zone_name: str
    :return: Zone name
    :rtype: str
    """
    i, r_zone_name = 0, zone_name
    while fabric_obj.r_zone_obj(r_zone_name) is not None:
        r_zone_name = zone_name + '_' + str(i)

    return r_zone_name


def _unique_zonecfg_name(fabric_obj, zonecfg_name):
    """Ensure that a zone name is unique by appending a number if alias_name is already used

    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :param zonecfg_name: Alias name
    :type zonecfg_name: str
    :return: Zone name
    :rtype: str
    """
    i, r_zonecfg_name = 0, zonecfg_name
    while fabric_obj.r_zonecfg_obj(r_zonecfg_name) is not None:
        r_zonecfg_name = zonecfg_name + '_' + str(i)

    return r_zonecfg_name


def _zone_option_all_func(fabric_obj, zonecfg, rnid_d, notes_l):
    """Returns a list with a single zone dictionary that includes all ports except ICL ports

    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :param zonecfg: Zone configuration name
    :type zonecfg: str
    :param rnid_d: RNID dictionary. Output of _rnid_groups()
    :type rnid_d: dict
    :param notes_l: List of notes to add to the notes page
    :type notes_l: list
    :return: Dictionaries with zone DB definitions for use with brcddb.brcddb_zone.add_zone_worksheet()
    :rtype: list
    """
    buf = brcddb_fabric.best_fab_name(fabric_obj, fid=True)
    zdb_l = [dict(Zone_Object='comment', Comments='All ports except ICL ports. Fabric: ' + buf)]
    member_l, comment_l = list(), list()
    zone_name = _unique_zone_name(fabric_obj, 'FICON_zone')

    for switch_obj in fabric_obj.r_switch_objects():
        did = str(switch_obj.r_did())
        for port_obj in [obj for obj in switch_obj.r_port_objects() if not obj.r_is_icl()]:
            member_l.append(did + ',' + str(port_obj.r_index()))
            buf = brcddb_port.port_best_desc(port_obj) + ', Switch: '
            buf += brcddb_switch.best_switch_name(port_obj.r_switch_obj(), fid=True, did=True)
            buf += ', Port: ' + port_obj.r_obj_key()
            comment_l.append(buf)
    if len(member_l) > 1:
        zdb_l.append(dict(Zone_Object='zone', Action='create', Name=zone_name, Member=member_l, Comments=comment_l))
        zdb_l.append(dict(Zone_Object='zone_cfg', Action='create', Name=zonecfg, Member=[zone_name]))
    else:
        notes_l.append('Fabric ' + brcddb_fabric.best_fab_name(fabric_obj, fid=True) + ' contains fewer than 2 ports.')

    return zdb_l


def _zone_option_all_f_func(fabric_obj, zonecfg, rnid_d, notes_l):
    """Returns a zone dictionary with a single zone that includes all F ports. See _zone_option_all_func"""
    buf = brcddb_fabric.best_fab_name(fabric_obj, fid=True)
    zdb_l = [dict(Zone_Object='comment', Comments='All F-Ports. Fabric: ' + buf)]
    member_l, comment_l = list(), list()
    zone_name = _unique_zone_name(fabric_obj, 'FICON_zone')

    for switch_obj in fabric_obj.r_switch_objects():
        did = str(switch_obj.r_did())
        for port_obj in [obj for obj in switch_obj.r_port_objects() if obj.c_login_type().lower() == 'f-port']:
            member_l.append(did + ',' + str(port_obj.r_index()))
            buf = brcddb_port.port_best_desc(port_obj) + ', Switch: '
            buf += brcddb_switch.best_switch_name(port_obj.r_switch_obj(), fid=True, did=True)
            buf += ', Port: ' + port_obj.r_obj_key()
            comment_l.append(buf)
    if len(member_l) > 1:
        zdb_l.append(dict(Zone_Object='zone', Action='create', Name=zone_name, Member=member_l, Comments=comment_l))
        zdb_l.append(dict(Zone_Object='zone_cfg', Action='create', Name=zonecfg, Member=[zone_name]))
    else:
        notes_l.append('No RNID data found attached to fabric ' + brcddb_fabric.best_fab_name(fabric_obj, fid=True))

    return zdb_l


def _zone_option_dev_func(fabric_obj, zonecfg, rnid_d, notes_l):
    """Returns a list of zones for each CHPID with access to all control units. See _zone_option_all_func"""
    buf = brcddb_fabric.best_fab_name(fabric_obj, fid=True)
    zdb_l = [dict(Zone_Object='comment', Comments='Single zone for each CHPID for every control unit. Fabric: ' + buf)]
    temp_notes_l, zone_l = list(), list()

    # Create an alias for each control unit
    cu_alias_l = list()
    for key in ('DASD', 'Tape', 'CUP', 'CTC', 'Switch', 'IDG', 'Test', 'UNKN'):
        for sub_key, port_obj_l in rnid_d.get(key, dict()).items():
            member_l, comment_l = list(), list()
            alias = _unique_alias_name(fabric_obj, key + '_' + sub_key + '_alias')
            cu_alias_l.append(alias)
            for port_obj in port_obj_l:
                member_l.append(str(port_obj.r_did()) + ',' + str(port_obj.r_index()))
                buf = brcddb_port.port_best_desc(port_obj) + ', Switch: '
                buf += brcddb_switch.best_switch_name(port_obj.r_switch_obj(), fid=True, did=True)
                buf += ', Port: ' + port_obj.r_obj_key()
                comment_l.append(buf)
            zdb_l.append(dict(Zone_Object='alias', Action='create', Name=alias, Member=member_l, Comments=comment_l))

    # Create a zone for each CHPID
    for sub_key, port_obj_l in rnid_d.get('CPU', dict()).items():
        for port_obj in port_obj_l:
            tag = port_obj.r_get('rnid/tag')
            if tag is None:
                buf = 'Missing tag for switch ' + brcddb_switch.best_switch_name(port_obj.r_switch_obj()) + ', port '
                buf += port_obj.r_obj_key()
                temp_notes_l.append(buf)
            else:
                for alias in cu_alias_l:
                    zone_name = 'CHPID_' + sub_key + '_' + tag.replace('0x', '') + '_' + alias.split('_')[1]
                    zone_name = _unique_zone_name(fabric_obj, zone_name)
                    zone_l.append(zone_name)
                    zdb_l.append({
                        'Zone_Object': 'peer_zone',
                        'Action': 'create',
                        'Name': zone_name,
                        'Member': [str(port_obj.r_did()) + ',' + str(port_obj.r_index())],
                        'Principal Member': [alias]
                    })

    # Create the zone configuration
    if len(zone_l) == 0:
        temp_notes_l.append('Nothing to zone for "By_Device"')
    else:
        if len(temp_notes_l) > 0:
            zdb_l.insert(1, dict(Zone_Object='comment', Comments=temp_notes_l))
        zdb_l.append(dict(Zone_Object='zone_cfg', Action='create', Name=zonecfg, Member=zone_l))
    notes_l.extend(temp_notes_l)

    return zdb_l


def _zone_option_channel_func(fabric_obj, zonecfg, rnid_d, notes_l):
    """Returns a zone dictionaries specific for each channel and associated link addresses. See _zone_option_all_func"""
    buf = 'Zone by CHPID with access to defined link addresses only. Fabric: ' + \
          brcddb_fabric.best_fab_name(fabric_obj, fid=True)
    zdb_l = [dict(Zone_Object='comment', Comments=buf)]
    temp_notes_l, zone_l = list(), list()

    # Find all the CHPIDs in this fabric and create a CHPID zone that includes the link addresses
    for iocp_obj in fabric_obj.r_project_obj().r_iocp_objects():
        for chpid_obj in iocp_obj.r_path_objects():
            chpid_port_obj = brcddb_iocp.find_chpid(fabric_obj, iocp_obj.r_obj_key(), chpid_obj.r_obj_key())
            if chpid_port_obj is not None:
                comment_l, cu_di_l = list(), list()
                for addr in [
                    gen_util.pad_string(brcddb_iocp.link_addr_to_fc_addr(a, did=chpid_port_obj.r_did()), 6, '0')
                    for a in chpid_obj.r_link_addresses()
                ]:
                    cu_port_obj = fabric_obj.r_port_obj_for_pid(addr)
                    if cu_port_obj is None:
                        buf = 'Could not find address ' + addr.upper() + ' in fabric '
                        buf += brcddb_fabric.best_fab_name(fabric_obj, fid=True)
                        temp_notes_l.append(buf)
                    else:
                        cu_di_l.append(str(cu_port_obj.r_did()) + ',' + str(cu_port_obj.r_index()))
                        buf = brcddb_port.port_best_desc(cu_port_obj) + ', Switch: '
                        buf += brcddb_switch.best_switch_name(cu_port_obj.r_switch_obj(), fid=True, did=True)
                        buf += ', Port: ' + cu_port_obj.r_obj_key()
                        comment_l.append(buf)
                zone_name = _unique_zone_name(fabric_obj, 'CHPID_' + iocp_obj.r_obj_key() + '_' + chpid_obj.r_obj_key())
                zone_l.append(zone_name)
                zdb_l.append({
                        'Zone_Object': 'peer_zone',
                        'Action': 'create',
                        'Name': zone_name,
                        'Member': [str(chpid_port_obj.r_did()) + ',' + str(chpid_port_obj.r_index())],
                        'Principal Member': cu_di_l,
                        'Comments': comment_l,
                    })

    # Create the zone configuration
    if len(zone_l) == 0:
        notes_l.append('Nothing to zone for "By_Channel"')
    else:
        if len(temp_notes_l) > 0:
            zdb_l.insert(1, dict(Zone_Object='comment', Comments=notes_l))
        zdb_l.append(dict(Zone_Object='zone_cfg', Action='create', Name=zonecfg, Member=zone_l))
    notes_l.extend(temp_notes_l)

    return zdb_l


# _zone_options_l: List of dictionaries as follows:
# h     Help message used with _eh_l
# a     Pointer to the function that builds the report
# sname Sheet name
_zone_options_l = (
    dict(
        h=dict(p1='A single large standard zone with all physical ports in the fabric except ICL ports is created. '
                  'This is typical of zoning methods currently being used for mainframe environments.'),
        a=_zone_option_all_func,
        sname='All_Ports',
    ),
    dict(
        h=dict(p1='The same as "All_Ports" except only online ports that are not E-Ports or ports configured for test '
                  'or loopback. Any port associated with a channel or control unit interface that was offline when the '
                  'data was collected will need to be added.'),
        a=_zone_option_all_f_func,
        sname='All_F_Ports',
    ),
    dict(
        h=[
            dict(p1='Device zoning creates a group of control units by serial number. A separate peer zone for each '
                    'CHPID and each group of control units is created.'),
            dict(p1=''),
            dict(p1='There is no need for mainframe storage administrators to know the difference between a zone and a '
                    'peer zone, but for the curious:'),
            dict(p1='', b1_prefix='  ', b1_char='*'),
            dict(b1='Peer zones have two membership lists, member and principal member.'),
            dict(b1='Any port in the member list can communicate with any port in the principal list.'),
            dict(b1='Ports in the member list cannot communicate with other ports in the member list. If this is '
                    'desired, you need to create another zone that permits it.'),
            dict(b1='Ports in the principal list cannot communicate with other ports in the principal list.'),
            dict(p1=''),
            dict(p1='In addition to creating smaller zones, using peer zones is ideal because it ensures that low '
                    'level fibre channel traffic from control units is not presented to other control unit ports.'),
            dict(p1=''),
            dict(p1='**WARNING** Any storage port that presents RNID data is considered a control unit. Mainframe '
                    'storage ports configured for peer-to-peer may or may not present RNID data. If RNID data is '
                    'presented, a zone for a CHPID to it will be created. Since there will not be any link '
                    'addresses defined to the port, that does not create a problem; however, the storage ports will '
                    'not be able to communicate with each other. You will need to add separate zones for disk '
                    'mirroring ports. Since disk mirroring ports do not behave as mainframe ports, consider putting '
                    'mirroring ports in their own logical fabric.'),
            dict(p1=''),
            dict(p1='Any channel or control unit interface that was offline when the data was collected will need to '
                    'be added.'),
            dict(p1=''),
            dict(p1='This is a little more flexible than "By_Channel" because supplying IOCPs is not required and a '
                    'new IODF can be activated with new link addresses for existing CHPIDs and control units. If there '
                    'are new CHPIDs, new devices, or any channel or device connection was moved, the zone database '
                    'will need to be updated accordingly. This is because FICON zoning is based on physical ports. New '
                    'channels, devices, or any relocation of channels or devices will result in a new physical port.'),
        ],
        a=_zone_option_dev_func,
        sname='By_Device'
    ),
    dict(
        h=[
            dict(p1='As with "By_Device", individual zones are built on a per channel basis. The difference is that '
                 'the channel zones only include ports associated with the "LINK=" statements in the CNTLUNIT macros. '
                 'Since the ports are determined from the "LINK=" statements, the control units do not have to be '
                 'online.'),
            dict(p1=''),
            dict(p1='Entry link addresses for channels are not specified in the IOCP files. They can only be '
                    'determined when the channel is online and the associated CHPID has been configured online.'),
            dict(p1=''),
            dict(p1='This method produces the smallest zones, but a zone configuration update will be required if:'),
            dict(p1='', b1_char='*', b1_prefix='  '),
            dict(b1='A new IODF is activated'),
            dict(b1='A previously offline CHPID is configured online'),
        ],
        a=_zone_option_channel_func,
        sname='By_Channel'
    ),
)

_eh_l = [
    dict(p1='**Overview**'),
    dict(p1=''),
    dict(p1='The purpose of this script is to provide a simple method to create zones for mainframe environments when '
            'SANnav is not used or when alternative zoning methods are desired. The output is a configuration workbook '
            'that can be used as input to zone_config.py.'),
    dict(p1=''),
    dict(p1='Three step approach:'),
    dict(p1=''),
    dict(b1_char='1'),
    dict(b1='Collect data using multi_capture.py or capture.py.'),
    dict(b1='Use this script to create a zone configuration workbook.'),
    dict(b1='Use zone_config.py to implement the zone configuration.'),
    dict(p1=''),
    dict(p1='Some of the zoning method options are based on connections to the fabric. Since its often not possible to '
            'have a fully configured fabric with all intended connections online, this approach allows you to modify '
            'the zone definitions to account for channels and devices not connected at the time data from switches '
            'was collected.'),
    dict(p1=''),
    dict(p1='Since the zone_config.py script has an option to output CLI, this also makes it easy for service '
            'organizations to create CLI zone scripts for SAN implementation engineers that do not have access to a '
            'Python interpreter with the brcddb and brcdapi libraries installed.'),
    dict(p1=''),
    dict(p1='Zone configurations are based on a fabric. This is why the -wwn parameter is required. Use the -scan '
            'parameter to determine the fabric world wide name (WWN). The fabric WWN is not needed for anything other '
            'than identifying which fabric in the data collection, specified with the -i parameter. to base the zone '
            'configuration on.'),
    dict(p1=''),
    dict(p1='**Naming Conventions**'),
    dict(p1='', b1_char='*', b1_prefix='  ', b2_char='-', b2_prefix='      '),
    dict(b1='A unique zone configuration name, -cfg, is enforced.'),
    dict(b1='Where applicable, zone name and alias names are a hash of the serial number, CHPID tag, and device type.'),
    dict(b1='Creating unique zone database items make it easy to:'),
    dict(b2='Back out zoning changes.'),
    dict(b2='Clean out previous zone databases by simply running the zone_config.py script with "full_purge"'),
    dict(p1=''),
    dict(p1='**Output File**'),
    dict(p1=''),
    dict(p1='There is a sheet for each zoning method. If IOCPs were not provided, the "By_Channel" work sheet will be '
            'empty. See "Zoning Methods" for details.'),
    dict(p1=''),
    dict(p1='When creating a zone object, If the zone object name already exists, a new one is created with a unique '
            'number appended to ensure it has a unique name. This makes it easy to use the "full_purge" action in '
            'zone_config.py workbooks to clean out no longer used zone database items after the zone configuration is '
            'validated.'),
    dict(p1=''),
    dict(p1='**Zoning Tips**'),
    dict(p1=''),
    dict(p1='Create a new zone configuration whenever modifying the zone database of an existing fabric so that your '
            'back-out plan is to simply activate the old zone configuration.'),
    dict(p1=''),
    dict(p1='**Background**'),
    dict(p1=''),
    dict(p1='In mainframe environments, controlling what servers (channels) can access what devices (control units) is '
            'handled by the IODF. In open systems environments, this function is handled with fabric zoning in the SAN '
            'switches. The most common approach to zoning for mainframes has been to create a single large zone. These '
            'large zones have become somewhat of a concern, so in addition to creating one large zone, this script '
            'also provides alternate zoning methods intended to reduce zone size.'),
    dict(p1=''),
    dict(p1='**Zoning Methods**'),
]
# Add the Notes page, including the zone options. While in this loop, we'll add the zone options to the extended help
_general_notes_l = [
    [dict(t='Tab', font=_bold_font), dict(t='Description', font=_bold_font)]
]
for _d in _zone_options_l:
    _eh_l.extend([dict(p1=''), dict(p1='**' + _d['sname'] + '**'), dict(p1=''),])
    _buf_l = gen_util.convert_to_list(_d['h'])
    _eh_l.extend(_buf_l)
    _buf_l = gen_util.format_text(_buf_l)
    _general_notes_l.append([dict(t=_d['sname']), dict(t=None if len(_buf_l) == 0 else _buf_l[0])])
    if len(_buf_l) > 1:
        for _buf in _buf_l[1:]:
            _general_notes_l.append([dict(), dict(t=_buf)])
_eh_l.append(dict(p1=''))

# Input parameter definitions
_input_d = dict(
    i=dict(r=False,
           h='Required unless using -eh. Name of input file generated by capture.py, combine.py, or multi_capture.py. '
             'Extension ".json" is automatically added.'),
    o=dict(r=False,
           h='Required unless using -eh or -scan. Output file file name. Extension ".xlsx" is automatically added.'),
    wwn=dict(r=False, h='Required unless using -eh or -scan. Fabric WWN. Run with -scan option if unknown.'),
    cfg=dict(r=False,
             h='Required unless using -eh or -scan. This name of the new zone configuration to create.'),
    iocp=dict(r=False,
              h='Required when the zoning type, -z, is "channel". Name of folder with IOCP files. All files in this '
                'folder must be IOCP files (build I/O configuration statements from HCD) and must begin with the CEC '
                'serial number followed by \'_\'. Leading 0s are not required. Example, for a CPC with serial number '
                '12345: 12345_M90_iocp.txt. Generates zone configurations based on channel paths.'),
)
_input_d.update(gen_util.parseargs_scan_d.copy())
_input_d.update(gen_util.parseargs_eh_d.copy())
_input_d.update(gen_util.parseargs_log_d.copy())


def _rnid_groups(fabric_obj):
    """Sorts out the mainframe attachments by control unit and channel

    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :return: Dictionary: key - generic device type. Value - dictionary, key - S/N, value - list of port objects
    :rtype: dict
    """
    # Sort out the ports by sequence number (S/N) and generic type.
    rnid_d = dict()
    for port_obj in [obj for obj in fabric_obj.r_port_objects() if obj.r_is_online()]:
        port_rnid_d = port_obj.r_get('rnid')
        if isinstance(port_rnid_d, dict):
            generic_type = brcddb_iocp.generic_device_type(port_rnid_d.get('type-number'))
            port_rnid_seq = gen_util.remove_leading_char(port_rnid_d.get('sequence-number', 'FFFFFFFFFFFF'), '0')

            # Get the generic device type dictionary from rnid_d. If it hasn't been added yet, add it.
            generic_type_d = rnid_d.get(generic_type)
            if not isinstance(generic_type_d, dict):
                generic_type_d = dict()
                rnid_d.update({generic_type: generic_type_d})

            # Get the port object list for this sequence number
            port_obj_l = generic_type_d.get(port_rnid_seq)
            if not isinstance(port_obj_l, list):
                port_obj_l = list()
                generic_type_d.update({port_rnid_seq: port_obj_l})

            # Add the port object
            port_obj_l.append(port_obj)

    return rnid_d


def pseudo_main(fabric_obj, output_file, zonecfg, iocp_l):
    """Basically the main(). Did it this way so that it can easily be used as a standalone module or called externally.

    :param fabric_obj: Fabric object
    :type fabric_obj: brcddb.classes.fabric.FabricObj
    :param output_file: Name of zone configuration workbook
    :type output_file: str
    :param zonecfg: Zone configuration name
    :type zonecfg: str
    :param iocp_l: IOCP files to read
    :type iocp_l: list
    :return: Exit code. See exist codes in brcddb.brcddb_common
    :rtype: int
    """
    global _zone_options_l, _general_notes_l

    ec, notes_l = brcddb_common.EXIT_STATUS_OK, list()

    # Build cross-references and custom search terms.
    brcdapi_log.log('Building cross-references', echo=True)
    proj_obj = fabric_obj.r_project_obj()
    brcddb_project.build_xref(proj_obj)
    brcddb_project.add_custom_search_terms(proj_obj)
    for file in iocp_l:
        brcddb_iocp.parse_iocp(proj_obj, file)
    rnid_d = _rnid_groups(fabric_obj)

    # Create a workbook and add a worksheet for each zone method
    brcdapi_log.log('Creating the zone configuration workbook, ' + output_file, echo=True)
    wb = brcddb_zone.create_zone_workbook()
    # Insert the "About" worksheet
    buf_l = (
        'The contents of this Workbook were generated from a script that reads a data from a data collection file or '
        'directly from a switch. It also reads IOCP plain text files. It generates a workbook to be used with '
        'zone_config.py. The zone and alias naming convention is a hash of the device type, serial number, and tag.'
        '',
        'See the notes page for details regarding the different worksheets and naming conventions.',
    )
    report_utils.about_page(wb, 1, 'About', os.path.basename(__file__), __version__, buf_l)
    for zone_d in _zone_options_l:
        brcddb_zone.add_zone_worksheet(wb, zone_d['sname'], zone_d['a'](fabric_obj, zonecfg, rnid_d, notes_l), 1)

    # Add the notes sheet
    sheet = wb.create_sheet(index=1, title='Notes')
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    for col in range(0, len(_zonecfg_col)):
        sheet.column_dimensions[xl.get_column_letter(col+1)].width = _zonecfg_col[col]

    # Add the content to the Notes sheet
    row = 1
    for buf in notes_l:
        excel_util.cell_update(sheet, row, 1, buf, font=_std_font, align=_align_wrap)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1
    if len(notes_l) > 0:
        row += 1
    for row_l in _general_notes_l:
        col = 1
        for col_d in row_l:
            excel_util.cell_update(
                sheet,
                row,
                col,
                col_d.get('t'),
                font=col_d.get('font', _std_font),
                align=col_d.get('align', _align_wrap),
                border=col_d.get('border'),
                link=col_d.get('link')
            )
            col_span = col_d.get('span', 1)
            if col_span > 1:
                sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + col_span-1)
            col += col_span
        row += 1

    # Save the workbook
    try:
        brcdapi_log.log('Saving: ' + output_file, echo=True)
        excel_util.save_report(wb, output_file, 'Consoli-Solutions')
    except (FileExistsError, FileNotFoundError):
        brcdapi_log.log('A folder in ' + output_file + ' does not exist.', echo=True)
        ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
    except PermissionError:
        brcdapi_log.log('Permission error writing ' + output_file + '. This usually happens when the file is open.',
                        echo=True)
        ec = brcddb_common.EXIT_STATUS_INPUT_ERROR

    return ec


def _get_input():
    """Parses the module load command line

    :return: Exit code. See exist codes in brcddb.brcddb_common
    :rtype: int
    """
    global __version__, _input_d, _eh_l
    
    ec = brcddb_common.EXIT_STATUS_OK

    # Get command line input
    args_d = gen_util.get_input(
        'Creates workbooks for use with zone_config.py for common mainframe zoning operations.',
        _input_d
    )

    # Set up logging
    brcdapi_log.open_log(
        folder=args_d['log'],
        suppress=args_d['sup'],
        no_log=args_d['nl'],
        version_d=brcdapi_util.get_import_modules()
    )

    # Setup feedback messages for options
    error_msg = '**ERROR**: Missing required input parameter.'
    feedback_d = dict(
        i=dict(h=str(args_d['i']), rl=('eh',), e=error_msg),
        o=dict(h=str(args_d['o']), rl=('eh', 'scan'), e=error_msg),
        wwn=dict(h=str(args_d['wwn']), rl=('eh', 'scan'), e=error_msg),
        cfg=dict(h=str(args_d['cfg']), rl=('eh', 'scan'), e=error_msg),
    )
    for key, d in feedback_d.items():
        try:
            for required_key in d['rl']:
                if args_d[required_key]:
                    raise Found  # The parameter is not required: -eh or -scan
            if args_d[key] is None:
                d['h'] += ' ' + d['e']
                ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
        except Found:
            continue

    # Read the project file, -i
    proj_obj, fabric_obj = None, None
    if args_d['i'] is not None and not args_d['eh']:
        in_file = brcdapi_file.full_file_name(args_d['i'], '.json')
        try:
            proj_obj = brcddb_project.read_from(in_file)
            if proj_obj is None:  # Error messages are sent to the log in brcddb_project.read_from() if proj_obj is None
                feedback_d['i'] += ' Error reading file. Refer to previous messages for details.'
                ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
            elif args_d['wwn'] is not None:
                fabric_obj = proj_obj.r_fabric_obj(args_d['wwn'])
                if fabric_obj is None:
                    feedback_d['wwn']['h'] += ' **ERROR** Does not exist in ' + args_d['i']
                    ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
        except FileNotFoundError:
            feedback_d['i']['h'] += ' **ERROR**: File not found.'
            ec = brcddb_common.EXIT_STATUS_INPUT_ERROR
        except FileExistsError:
            feedback_d['i']['h'] += ' **ERROR**: One of the folders does not exist.'
            ec = brcddb_common.EXIT_STATUS_INPUT_ERROR

    # Read the directory of IOCP files, -iocp
    iocp_l = list()
    if args_d['iocp'] is not None:
        iocp_l = brcdapi_file.read_directory(args_d['iocp'], full=True)
        if len(iocp_l) == 0:
            feedback_d['iocp']['h'] += ' **ERROR**: Folder not found or folder is empty.'
            ec = brcddb_common.EXIT_STATUS_INPUT_ERROR

    # Make sure the zone configuration name is unique
    if fabric_obj is not None and isinstance(args_d['cfg'], str):
        zonecfg = _unique_zonecfg_name(fabric_obj, args_d['cfg'])
        if zonecfg != args_d['cfg']:
            feedback_d['cfg']['h'] = ' **ERROR**:' + args_d['cfg'] + ' already used. Suggestion: ' + zonecfg
            ec = brcddb_common.EXIT_STATUS_INPUT_ERROR

    # Command line feedback
    ml = ['',
          os.path.basename(__file__) + ', ' + __version__,
          'Input file, -i:           ' + feedback_d['i']['h'],
          'Output file, -o:          ' + feedback_d['o']['h'],
          'Fabric WWN, -wwn:         ' + feedback_d['wwn']['h'],
          'Zone configuration, -cfg: ' + feedback_d['cfg']['h'],
          'IOCP, -iocp:              ' + str(args_d['iocp']),
          'Extended help, -eh:       ' + str(args_d['eh']),
          'Scan, -scan:              ' + str(args_d['scan']),
          'Log, -log:                ' + str(args_d['log']),
          'No log, -nl:              ' + str(args_d['nl']),
          'Suppress, -sup:           ' + str(args_d['sup']),
          '',]
    brcdapi_log.log(ml, echo=True)

    # Extended help, Scan, and error bail out.
    if args_d['eh']:
        brcdapi_log.log(gen_util.format_text(_eh_l), echo=True)
        return brcddb_common.EXIT_STATUS_INPUT_ERROR
    if args_d['scan'] and proj_obj is not None:
        brcdapi_log.log(brcddb_project.scan(proj_obj, fab_only=False, logical_switch=True), echo=True)
        return brcddb_common.EXIT_STATUS_INPUT_ERROR
    if not isinstance(args_d['i'], str) or not isinstance(args_d['o'], str):
        return brcddb_common.EXIT_STATUS_INPUT_ERROR

    return ec if ec != brcddb_common.EXIT_STATUS_OK else \
        pseudo_main(fabric_obj, brcdapi_file.full_file_name(args_d['o'], '.xlsx'), args_d['cfg'], iocp_l)


##################################################################
#
#                    Main Entry Point
#
###################################################################
if _DOC_STRING:
    print('_DOC_STRING is True. No processing')
    exit(0)

if _STAND_ALONE:
    _ec = _get_input()
    brcdapi_log.close_log(['', 'Processing Complete. Exit code: ' + str(_ec)], echo=True)
    exit(_ec)
