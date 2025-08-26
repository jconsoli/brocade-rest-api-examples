#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Copyright 2023, 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

**License**

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

**Description**

Parses IOCP files and generates planning workbooks

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Fixed error message when there is an error adding a sheet.                            |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 03 Apr 2024   | Added version numbers of imported libraries.                                          |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 06 Dec 2024   | Fixed options for port naming conventions.                                            |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 01 Mar 2025   | Error message enhancements.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.5     | 25 Aug 2025   | Use brcddb.util.util.get_import_modules to dynamically determined imported libraries. |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2024, 2025 Consoli Solutions, LLC'
__date__ = '25 Aug 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.5'

import datetime
import sys
import os
import collections
import openpyxl as xl_wb
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.excel_util as excel_util
import brcdapi.excel_fonts as excel_fonts
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util
import brcdapi.file as brcdapi_file
import brcddb.brcddb_project as brcddb_project
import brcddb.brcddb_common as brcddb_common
import brcddb.util.iocp as brcddb_iocp

_DOC_STRING = False  # Should always be False. Prohibits any code execution. Only useful for building documentation

_config_d = dict(x64='templates/X6-4_Switch_Configuration.xlsx',
                 x68='templates/X6-8_Switch_Configuration.xlsx',
                 x74='templates/X7-4_Switch_Configuration.xlsx',
                 x78='templates/X7-8_Switch_Configuration.xlsx',
                 fixed='templates/Fixed_Port_Switch_Configuration.xlsx')
_valid_switch_type_l = [str(k) for k in _config_d.keys()]

_buf = 'Optional. Switch type. Valid switch types are: ' + ', '.join(_valid_switch_type_l) + '. The default is x78'
_input_d = dict(
    p=dict(h='Required. Prefix for the name of the switch configuration workbooks. The name of this workbook begins '
             'with this prefix followed by an "_", the switch DID and ".xlsx". A file path may be embedded in the '
             'prefix.'),
    t=dict(r=False, d='x78', v=_valid_switch_type_l, h=_buf),
    iocp=dict(h='Required. Name of folder containing IOCP files to be parsed.'),
    map=dict(r=False,
             h='Optional. A map of Switch IDs to domain Both values must be in hex. Separate multiple pairs with a '
               '";". For example, to map Switch ID F0 to DID 32 (0x20) and Switch ID F1 to DID 33 (0x21): '
               '"F0,20;F1,21".'),
    d=dict(r=False, d=False, t='bool',
           h='Optional. No parameters. Instead of writing the workbooks, plain text files of the IOCP files filtered '
             'for lines containing "CNTLUNIT" and "LINK=".')
)
_input_d.update(gen_util.parseargs_log_d.copy())

# Common worksheet cell formatting
_hdr1_font = excel_fonts.font_type('hdr_1')
_hdr2_font = excel_fonts.font_type('hdr_2')
_bold_font = excel_fonts.font_type('bold')
_white_bold_font = excel_fonts.font_type('white_bold')
_std_font = excel_fonts.font_type('std')
_config_slot = excel_fonts.fill_type('config_slot')
_border = excel_fonts.border_type('thin')
_alignment = excel_fonts.align_type('wrap')
_center_align = excel_fonts.align_type('wrap_center')

_skip_sheets = ('CLI_Bind',)
_copy_sheets_d = collections.OrderedDict()
_copy_sheets_d['VC'] = dict(font=_std_font)
_copy_sheets_d['lists'] = dict(font=_std_font)
_copy_sheets_d['Chassis'] = dict(width_l=(25, 20, 80), font=_std_font, border=_border, align=_alignment)
_copy_sheets_d['Instructions'] = dict(width_l=(88,), font=_std_font, border=_border, align=_alignment)

_com_switch_type_d = dict(x64='x4', x74='x4', x68='x8', x78='x8', fixed='fixed')
_generic_blade_type_d = dict(x4={3: 'pc_48', 4: 'pc_48', 5: 'core_4', 6: 'core_4', 7: 'pc_48', 8: 'pc_48'},
                             x8={3: 'pc_48', 4: 'pc_48', 5: 'pc_48', 6: 'pc_48', 7: 'core_8', 8: 'core_8',
                                 9: 'pc_48', 10: 'pc_48', 11: 'pc_48', 12: 'pc_48'},
                             fixed={0: 'fixed'})

"""_fill_d: Row and column numbers are 0 referenced (so add 1 to match Excel). First key is the generic blade type,
Second key is the row number. Third key is the column number. The value is 0 for ASIC 0 and 1 for ASIC 1. This is used
for determining the fill color in parse_iocp.py"""
_fill_d = dict(
    pc_48={
        3: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
            13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 1, 22: 1, 23: 1},
        4: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
            13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 1, 22: 1, 23: 1},
        5: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
            13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 1, 22: 1, 23: 1},
        6: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
            13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 1, 22: 1, 23: 1},
        7: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
            13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 1, 22: 1, 23: 1},
        8: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
            13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 1, 22: 1, 23: 1},
        9: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
            13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 1, 22: 1, 23: 1},
        10: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 1, 14: 1, 15: 1, 16: 1, 17: 1, 18: 1, 19: 1, 20: 1, 21: 1, 22: 1, 23: 1},
        11: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        12: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        13: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        14: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        15: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        16: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        17: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        18: {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        19: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        20: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        21: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        22: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        23: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        24: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        25: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
        26: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0,
             13: 0, 14: 0, 15: 0, 16: 0, 17: 0, 18: 0, 19: 0, 20: 0, 21: 0, 22: 0, 23: 0},
    },
    core_4={
        3: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        4: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        5: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        6: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        7: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        8: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        9: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        10: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        11: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        12: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        13: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        14: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        15: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        16: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        17: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        18: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
    },
    core_8={
        3: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        4: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        5: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        6: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        7: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        8: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        9: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        10: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        11: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        12: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        13: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        14: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        15: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        16: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        17: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        18: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        19: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        20: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        21: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        22: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        23: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        24: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        25: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        26: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        27: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        28: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        29: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        30: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        31: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        32: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        33: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
        34: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 7: 1, 8: 1, 9: 1, 10: 1, 11: 1},
    },
    fixed={
        3: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        4: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        5: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        6: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        7: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        8: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        9: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        10: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        11: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        12: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        13: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        14: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        15: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        16: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        17: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        18: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        19: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        20: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        21: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        22: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        23: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        24: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        25: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        26: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        27: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        28: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        29: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        30: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        31: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        32: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        33: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        34: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        35: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        36: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        37: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        38: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        39: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        40: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        41: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        42: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        43: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        44: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        45: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        46: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        47: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        48: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        49: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        50: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        51: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        52: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        53: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        54: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        55: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        56: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        57: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        58: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        59: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        60: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        61: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        62: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        63: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        64: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        65: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        66: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        67: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        68: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        69: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        70: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        71: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        72: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        73: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        74: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        75: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        76: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        77: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        78: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        79: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        80: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        81: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        82: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        83: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        84: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        85: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        86: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        87: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        88: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        89: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        90: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        91: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        92: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        93: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        94: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        95: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        96: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        97: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        98: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        99: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        100: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        101: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        102: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        103: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        104: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        105: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        106: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        107: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        108: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        109: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        110: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        111: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        112: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        113: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        114: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        115: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        116: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        117: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        118: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        119: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        120: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        121: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        122: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        123: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        124: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        125: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        126: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        127: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        128: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        129: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
        130: {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0},
    }
)
_fill_asic_d = {0: excel_fonts.fill_type('config_asic_0'), 1: excel_fonts.fill_type('config_asic_1')}


class InvalidDID(Exception):
    pass


"""
The following are used in _row_val_x_l and _special_d. When generating the tables, they are used as follows:

+-----------+---------------------------------------+
| Mnemonic  | Description                           |
+===========+=======================================+
| $nr       | Next row                              |
+-----------+---------------------------------------+
| $r        | Current row                           |
+-----------+---------------------------------------+
| $did      | Domain ID in hex.                     |
+-----------+---------------------------------------+
| $tv       | Use template value                    |
+-----------+---------------------------------------+
"""
_link_addr_0 = '=IF(HEX2DEC(D$r)>15,IF(HEX2DEC(C$r) > 15,CONCATENATE(C$r,D$r),CONCATENATE("0",C$r,D$r)),' \
               'IF(HEX2DEC(C$r) > 15,CONCATENATE(C$r,"0",D$r),CONCATENATE("0",C$r,"0",D$r)))'
_link_addr_1 = '=IF(HEX2DEC(P$r)>15,IF(HEX2DEC(O$r) > 15,CONCATENATE(O$r,P$r),CONCATENATE("0",O$r,P$r)),' \
               'IF(HEX2DEC(O$r) > 15,CONCATENATE(O$r,"0",P$r),CONCATENATE("0",O$r,"0",P$r)))'

""" _row_val_d: The first key is the generic blade type. The dictionary for each blade type is as follows:

+-----------+---------------------------------------------------------------------------------------+
| Key       | Description                                                                           |
+===========+=======================================================================================+
| row       | A dictionary as follows:                                                              |
|           |   start   The first row (zero based) in the workbook template for the port details.   |
|           |   end     The last row (zero based) in the workbook template for the port details.    |
|           |   skip    Zero based rows to skip in the workbook template for the port details.      |
|           |   row_l   A list of rows to read from the workbook template for the port details.     |
+-----------+---------------------------------------------------------------------------------------+
| col       | Same as row but for the columns                                                       |
+-----------+---------------------------------------------------------------------------------------+
| ad        | If True, this column is the "Attached Device". Default is False.                      |
+-----------+---------------------------------------------------------------------------------------+
| did       | If True, this cell is the Domain ID.                                                  |
+-----------+---------------------------------------------------------------------------------------+
| port_addr | If True, this column is the Port Address. Default is False.                           |
+-----------+---------------------------------------------------------------------------------------+
| skip      | If True, skip adding this column to the worksheet. Default is False.                  |
+-----------+---------------------------------------------------------------------------------------+
| width     | Column width                                                                          |
+-----------+---------------------------------------------------------------------------------------+
| 1-x       | Zero based column numbers as strings.                                                 |
+-----------+---------------------------------------------------------------------------------------+
"""
_row_val_d = dict(  # First key is generic switch type as returned from Key is the zero based column number
    pc_48={
        'row': dict(start=3, end=27, row_l=list()),
        'col': dict(start=1, end=24, col_l=list()),
        0: dict(width=2, skip=True),
        1: dict(width=5, align=_center_align),  # Port
        2: dict(width=6, align=_center_align, val='=C$nr', did=True),  # DID
        3: dict(width=6, align=_center_align, val='=DEC2HEX(HEX2DEC(D$nr)+1)', port_addr=True),  # Port Addr
        4: dict(width=6, align=_center_align),  # Index
        5: dict(width=7, align=_center_align, val=_link_addr_0),  # Link Addr
        6: dict(width=5, align=_center_align, val='=G$nr'),  # FID
        7: dict(width=32, align=_alignment, ad=True),  # Attached Device
        8: dict(width=32, align=_alignment),  # Port Name
        9: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(F$r,4,1)),VC!A1:C16)'),  # Low
        10: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(F$r,4,1)),VC!A1:B16)'),  # Med
        11: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(F$r,4,1)),VC!A1:D16)'),  # High
        12: dict(width=2, skip=True),
        13: dict(width=5, align=_center_align),  # Port
        14: dict(width=6, align=_center_align, val='=O$nr', did=True),  # DID
        15: dict(width=6, align=_center_align, val='=DEC2HEX(HEX2DEC(p$nr)+1)', port_addr=True),  # Port Addr
        16: dict(width=6, align=_center_align),  # Index
        17: dict(width=7, align=_center_align, val=_link_addr_1),  # Link Addr
        18: dict(width=5, align=_center_align, val='=S$nr'),  # FID
        19: dict(width=32, align=_alignment, ad=True),  # Attached Device
        20: dict(width=32, align=_alignment),  # Port Name
        21: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(R$r,4,1)),VC!A1:C16)'),  # Low
        22: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(R$r,4,1)),VC!A1:B16)'),  # Med
        23: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(R$r,4,1)),VC!A1:D16)'),  # High
    },
    core_4={
        'row': dict(start=3, end=19, row_l=list()),
        'col': dict(start=1, end=12, col_l=list()),
        0: dict(width=2, skip=True),
        1: dict(width=5, align=_center_align),  # Port
        2: dict(width=6, align=_center_align),  # Index
        3: dict(width=5, align=_center_align, val='=D$nr'),  # FID
        4: dict(width=32, align=_alignment),  # ICL Description
        5: dict(width=32, align=_alignment),  # Port Name
        6: dict(width=2, skip=True),
        7: dict(width=5, align=_center_align),  # Port
        8: dict(width=6, align=_center_align),  # Index
        9: dict(width=5, align=_center_align, val='=J$nr'),  # FID
        10: dict(width=32, align=_alignment),  # ICL Description
        11: dict(width=32, align=_alignment),  # Port Name
    },
    core_8={
        'row': dict(start=3, end=35, row_l=list()),
        'col': dict(start=1, end=12, col_l=list()),
        0: dict(width=2, skip=True),
        1: dict(width=5, align=_center_align),  # Port
        2: dict(width=6, align=_center_align),  # Index
        3: dict(width=5, align=_center_align, val='=D$nr'),  # FID
        4: dict(width=32, align=_alignment),  # ICL Description
        5: dict(width=32, align=_alignment),  # Port Name
        6: dict(width=2, skip=True),
        7: dict(width=5, align=_center_align),  # Port
        8: dict(width=6, align=_center_align),  # Index
        9: dict(width=5, align=_center_align, val='=J$nr'),  # FID
        10: dict(width=32, align=_alignment),  # ICL Description
        11: dict(width=32, align=_alignment),  # Port Name
    },
    fixed={
        'row': dict(start=3, end=131, row_l=list()),
        'col': dict(start=1, end=12, col_l=list()),
        0: dict(width=2, skip=True),
        1: dict(width=5, align=_center_align),  # Port
        2: dict(width=6, align=_center_align, val='=C$nr', did=True),  # DID
        3: dict(width=6, align=_center_align, val='=DEC2HEX(HEX2DEC(D$nr)+1)', port_addr=True),  # Port Addr
        4: dict(width=6, align=_center_align),  # Index
        5: dict(width=7, align=_center_align, val=_link_addr_0),  # Link Addr
        6: dict(width=5, align=_center_align, val='=G$nr'),  # FID
        7: dict(width=32, align=_alignment, ad=True),  # Attached Device
        8: dict(width=32, align=_alignment),  # Port Name
        9: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(F$r,4,1)),VC!A1:C16)'),  # Low
        10: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(F$r,4,1)),VC!A1:B16)'),  # Med
        11: dict(width=5, align=_center_align, val='=LOOKUP(HEX2DEC(MID(F$r,4,1)),VC!A1:D16)'),  # High
    }
)
# Figure out what cells by row and column to read from the workbook template
for _k0, _d0 in _row_val_d.items():
    for _row in range(_d0['row']['start'], _d0['row']['end']):
        _d0['row']['row_l'].append(_row)
    for _k1, _d1 in _d0.items():
        if isinstance(_k1, int):
            _d0['col']['col_l'].append(_k1)

"""Normally, cell values are determined by _row_val_d. _row_val_d contains the common cell value which are either None,
read from the workbook template in _config_d, or a calculation based on other cells. Those cell references are either
on the same row or a previous row. _special_d handles special cases such the first DID which references a cell from the
previous sheet.

+-------+---------------------------------------------------------------+
| Key   | Description                                                   |
+=======+===============================================================+
| 0     | Generic switch type (see _com_switch_type_d)                  |
+-------+---------------------------------------------------------------+
| 1     | Generic blade type (see _generic_blade_type_d)                |
+-------+---------------------------------------------------------------+
| 2     | Slot number                                                   |
+-------+---------------------------------------------------------------+
| 3     | Zero based row number                                         |
+-------+---------------------------------------------------------------+
| 4     | Zero based column number                                      |
+-------+---------------------------------------------------------------+
"""
_special_d = dict(
    x4=dict(
        pc_48={3: {26: {2: '$did',  # DID
                        3: 0,  # Port Addr
                        6: 1,  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}},  # FID
               4: {26: {2: "='Slot 3'!O4",  # DID
                        3: "=DEC2HEX(HEX2DEC('Slot 3'!P4)+1)",  # Port Addr
                        6: "='Slot 3'!S4",  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}},  # FID
               5: {26: {3: 128, 9: '=D4'}},
               6: {26: {3: "='Slot 5'!J4", 9: '=D4'}},
               7: {26: {2: "='Slot 4'!O4",  # DID
                        3: "=DEC2HEX(HEX2DEC('Slot 4'!P4)+1)",  # Port Addr
                        6: "='Slot 4'!S4",  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}},  # FID
               8: {26: {2: "='Slot 7'!O4",  # DID
                        3: "=DEC2HEX(HEX2DEC('Slot 7'!P4)+1)",  # Port Addr
                        6: "='Slot 7'!S4",  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}}},  # FID
        core_4={5: {18: {3: 128, 9: '=D4'}},  # FID
                6: {18: {3: 'Slot 5!J4', 9: '=D4'}}},  # FID
        core_8={7: {34: {3: 128, 9: '=D4'}},  # FID
                8: {34: {3: 'Slot 5!J4', 9: '=D4'}}},  # FID
    ),
    x8=dict(
        pc_48={3: {26: {2: '$did',  # DID
                        3: 0,  # Port Addr
                        6: 1,  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}},  # FID
               4: {26: {2: "='Slot 3'!O4",  # DID
                        3: "=DEC2HEX(HEX2DEC('Slot 3'!P4)+1)",  # Port Addr
                        6: "='Slot 3'!S4",  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}},  # FID
               5: {26: {2: "='Slot 4'!O4",  # DID
                        3: "=DEC2HEX(HEX2DEC('Slot 4'!P4)+1)",  # Port Addr
                        6: "='Slot 4'!S4",  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}},  # FID
               6: {26: {2: "='Slot 5'!O4",  # DID
                        3: "=DEC2HEX(HEX2DEC('Slot 5'!P4)+1)",  # Port Addr
                        6: "='Slot 5'!S4",  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}},  # FID
               9: {26: {2: "='Slot 6'!O4",  # DID
                        3: "=DEC2HEX(HEX2DEC('Slot 6'!P4)+1)",  # Port Addr
                        6: "='Slot 6'!S4",  # FID
                        14: '=C4',  # DID
                        15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                        18: '=G4'}},  # FID
               10: {10: {2: '$tv',  # DID - This row is where 0xFF roles to 0x00 so take the DID from the template
                         3: '$tv',  # Port Addr - See note with DID
                         6: '$tv'},  # FID - See note with DID
                    26: {2: "='Slot 9'!O4",  # DID
                         3: "=DEC2HEX(HEX2DEC('Slot 9'!P4)+1)",  # Port Addr
                         6: "='Slot 9'!S4",  # FID
                         14: '=C4',  # DID
                         15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                         18: '=G4'}},  # FID
               11: {26: {2: "='Slot 10'!O4",  # DID
                         3: "=DEC2HEX(HEX2DEC('Slot 10'!P4)+1)",  # Port Addr
                         6: "='Slot 10'!S4",  # FID
                         14: '=C4',  # DID
                         15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                         18: '=G4'}},  # FID
               12: {26: {2: "='Slot 11'!O4",  # DID
                         3: "=DEC2HEX(HEX2DEC('Slot 11'!P4)+1)",  # Port Addr
                         6: "='Slot 11'!S4",  # FID
                         14: '=C4',  # DID
                         15: '=DEC2HEX(HEX2DEC(D4)+1)',  # Port Addr
                         18: '=G4'}}},  # FID
        core_4={7: {18: {3: 128, 9: '=D4'}},  # FID
                8: {18: {3: "='Slot 7'!J4", 9: '=D4'}}},  # FID
        core_8={7: {34: {3: 128, 9: '=D4'}},  # FID
                8: {34: {3: "='Slot 7'!J4", 9: '=D4'}}},  # FID
    ),
    fixed=dict(
        fixed={0: {130: {2: '$did',  # DID
                         3: 0,  # Port Addr
                         6: 1}}},  # FID
    )
)

# About worksheet
_about_sheet_l = (
    dict(f=_hdr1_font, t='Generated By'),
    None,
    dict(f=_std_font, t='parse_iocp.py'),
    dict(f=_std_font, t='Version: ' + __version__),
    None,
    dict(f=_hdr1_font, t='Disclaimer'),
    None,
    dict(f=_std_font, t='The contents of this Workbook were generated from a simple script that reads IOCPs (build I/O'
                        ' configuration statements from HCD). The script uses open source software covered by the GNU '
                        'General Public License or other open source license agreements. It does not consider all '
                        'supported IOCP syntax. It is only provided to give guidance in channel path planning.'),
    None,
    dict(f=_std_font, t='Information furnished by Broadcom is believed to be accurate and reliable. However, Broadcom '
                        'does not assume any liability arising out of the application or use of this information, nor '
                        'the application or use of any product or circuit described herein, neither does it convey any'
                        ' license under its patent rights nor the rights of others.'),
    None,
    dict(f=_hdr1_font, t='Description & Instructions'),
    None,
    dict(f=_std_font, t='The purpose of the parse_iocp script is to provide mainframe architects with a planning tool '
                        'to aid in determining physical switch ports for channel paths. Its typical use is for SAN '
                        'upgrades and migrations.'),
    None,
    dict(f=_std_font, t='The IOCPs are parsed and an Excel Workbook for each switch found in the IOCPs is generated. '
                        'The workbook contains worksheets matching the physical geometry of Brocade switches and '
                        'directors. Link addresses and control unit types are placed on worksheets such that the '
                        'physical port number and ASIC can easily be determined.'),
    None,
    dict(f=_std_font, t='Unless a conversion table of switch IDs to domain ID was provided, the SWITCH= is assumed'
                        ' to be the domain ID. This is necessary to determine:'),
    None,
    dict(f=_std_font, t='1) Which switch the CHPIDs are connected to.'),
    dict(f=_std_font, t='2) The full link address when single byte addressing is used.'),
    None,
    dict(f=_std_font, t='Summary:'),
    None,
    dict(f=_std_font, t='1) Parse all CHPID macros first'),
    dict(f=_std_font, t='2) Filter the list of CHPID macros to just those with SWITCH= statement'),
    dict(f=_std_font, t='3) Parse all CNTLUNIT macros'),
    dict(f=_std_font, t='4) Filter the list of CNTLUNIT macros to just those associated with switch CHPIDS'),
    dict(f=_std_font, t='5) Generates an Excel Workbook for each switch found with the following tabs:'),
    None,
    dict(f=_hdr2_font, t='About'),
    None,
    dict(f=_std_font, t='This tab'),
    None,
    dict(f=_hdr2_font, t='Summary'),
    None,
    dict(f=_std_font, t='Contains the following 3 summaries:'),
    None,
    dict(f=_std_font, t='A list of CHPIDs and the ports associated with the link addresses.'),
    dict(f=_std_font, t='A list of control units.'),
    dict(f=_std_font, t='A list of link addresses not supported by the switch geometry.'),
    None,
    dict(f=_hdr2_font, t='Switch_x'),
    None,
    dict(f=_std_font, t='This tab is intended to be filled in by the user. It is not used by this utility.  Read all '
                        'instructions with switch_config.py when using switch_config.py to configure switches from '
                        'this workbook.'),
    None,
    dict(f=_hdr2_font, t='Slot x'),
    None,
    dict(f=_std_font, t='For fixed port switches, “x” will always be 0. Otherwise, this is the slot number of the '
                        'director for the port card. The worksheet mimics the switch or port card geometry. Note that '
                        'with port cards, the addressing begins at the bottom of the card, just as it would be in an '
                        '8 slot director.'),
    None,
    dict(f=_std_font, t='To change the domain ID, manually set the cell in the “DID (Hex)” column to the desired DID. '
                        'The domain ID is automatically populated in all subsequent cells.'),
    dict(f=_std_font, t='To change the fabric ID, manually set the cell in the “FID” column to the desired FID. The '
                        'fabric ID is automatically populated in all subsequent cells.'),
    None,
    dict(f=_std_font, t='To change the port address (low byte of the link address), manually set the cell in the '
                        '“Port Addr (Hex)” column to the desired port address. The port address is automatically '
                        'populated in all subsequent cells. Note that the full link address, column “Link Addr”, is '
                        'automatically updated. WARNING: The workbook was built assuming sequential addressing '
                        'beginning with port address 0x00. If you change the port address, the prepopulated control '
                        'unit in the “Attached Device” column will no longer match the link address in the IOCP.'),
    None,
    dict(f=_std_font, t='The cells in the “Attached Device” column are pre-populated with the control unit type, '
                        'control unit number, and CHPIDs with link addresses defined for the associated control unit '
                        'found in the IOCPs. CHPIDs are not prepopulated because they do not have an entry link '
                        'address in HCD or the IOCP. The intent of this column is to provide planners with a starting '
                        'point and visual representation of how the switch is used. This column is not used by '
                        'switch_config.py. It is for documentation and planning purposes only.'),
    None,
    dict(f=_std_font, t='Do not change any other cells. All other cells are a calculation based on other cells.'),
    None,
    dict(f=_hdr1_font, t='Resources'),
    None,
    dict(f=_std_font, t='For FICON planning and switch configuration, Chapter 6 of the “IBM b-type Gen 7 '
                        'Installation, Migration, and Best Practices Guide”, sg248497, is applicable to both Gen 6 & '
                        'Gen 7.'),
    None,
    dict(f=excel_fonts.font_type('link'), t='https://www.redbooks.ibm.com/redpieces/abstracts/sg248497.html?Open'),
)

# Summary worksheet headers. First key is the column letter for the cell
_summary_sheet_d = dict(
    col=(10, 22, 98),
    chpid_d=dict(t='CHPIDs ($c)', h=('Tag', 'IOCP: CHPID', 'Ports Matching Link Addresses')),
    link_d=dict(t='Control Units ($c)', h=('Link Address', 'Type', 'CUNUM')),
    error_d=dict(t='Errors ($c)', h=('Area', 'Type', 'Description'))
)

# Configuration Worksheet
_switch_config_hdr = (dict(t='Area', c=20, align=_alignment),
                      dict(t='Parameter', c=20, align=_center_align),
                      dict(t='Comments', c=80, align=_alignment))
_dv_common_e = ' Please use pull down menu to pick an appropriate value.'  # Appended to all data validation error msgs
_dv_yn = DataValidation(type='list', formula1='"Yes,No"', allow_blank=False)
_dv_yn.errorTitle = 'Invalid Entry'
_dv_yn.error = 'Value must be "Yes" or "No".' + _dv_common_e
_dv_fid = DataValidation(type='list', formula1='lists!$B$2:$B$129', allow_blank=False)
_dv_fid.errorTitle = 'Invalid Entry'
_dv_fid.error = 'Invalid FID. FID must be an integer between 1-128.' + _dv_common_e
_dv_did = DataValidation(type='list', formula1='lists!$A$2:$A$240', allow_blank=False)
_dv_did.errorTitle = 'Invalid Entry'
_dv_did.error = 'Invalid DID. DID must be 0x01-0xEF.' + _dv_common_e
_dv_switch_type = DataValidation(type='list', formula1='"base,ficon,open"', allow_blank=False)
_dv_switch_type.errorTitle = 'Invalid Entry'
_dv_switch_type.error = 'Invalid switch type. Switch type must be base, ficon, or open' + _dv_common_e
_dv_fab_principal = DataValidation(type='list', formula1='lists!$D$2:$D$255', allow_blank=False)
_dv_fab_principal.errorTitle = 'Invalid Entry'
_dv_fab_principal.error = 'Invalid fabric priority. See FOS CLI fabricprincipal -priority for details.' + _dv_common_e
_dv_routing_policy = DataValidation(type='list', formula1='"default,DBR,EBR"', allow_blank=False)
_dv_routing_policy.errorTitle = 'Invalid Entry'
_dv_routing_policy.error = 'Invalid Routing Policy. Must be default, DBR, or EBR.' + _dv_common_e
_dv_port_name = DataValidation(type='list', formula1='"off,default,fdmi,dynamic,open -n,ficon -n"', allow_blank=False)
_dv_port_name.errorTitle = 'Invalid Entry'
_dv_port_name.error = 'Invalid Port Name. Must be off, default, fdmi, dynamic, open -n, or ficon -n.' + _dv_common_e
_dv_dup_wwn = DataValidation(type='list', formula1='"0,1,2"', allow_blank=False)
_dv_dup_wwn.errorTitle = 'Invalid Entry'
_dv_dup_wwn.error = 'Invalid duplicate WWN value. Value must be 0, 1, or 2.' + _dv_common_e
# _switch_config_d is used to determine the data validation for the "Parameter" column of the workbook. If it's not in
# _switch_config_d, the cell is assumed to be free form text (no data validation).
_switch_config_d = {
    'Fabric ID (FID)': dict(p=1, dv=_dv_fid),
    'Domain ID (DID)': dict(p='0x01', dv=_dv_did),
    'Insistent DID': dict(p='Yes', dv=_dv_yn),
    'Fabric Principal Enable': dict(p='No', dv=_dv_yn),
    'Fabric Principal Priority': dict(p='0x0A', dv=_dv_fab_principal),
    'Allow XISL': dict(p='No', dv=_dv_yn),
    'Enable Switch': dict(p='Yes', dv=_dv_yn),
    'Enable Ports': dict(p='No', dv=_dv_yn),
    'Switch Type': dict(p='ficon', dv=_dv_switch_type),
    'Duplicate WWN': dict(p=0, dv=_dv_dup_wwn),
    'Bind': dict(p='Yes', dv=_dv_yn),
    'Routing Policy': dict(p='default', dv=_dv_routing_policy),
    'Port Name': dict(p='None', dv=_dv_port_name),
    'Enable CUP': dict(p='Yes', dv=_dv_yn),
}

"""
+-------+-------+---------------------------------------------------------------------------------------------------+
| Key   | type  | Description                                                                                       |
+=======+=======+===================================================================================================+
| desc  | str   | Column descriptor.                                                                                |
+-------+-------+---------------------------------------------------------------------------------------------------+
| i     | int   | Index into key_d[desc][col_l]. Note that some boards have 2 columns so this distinguishes which   |
|       |       | column.                                                                                           |
+-------+-------+---------------------------------------------------------------------------------------------------+
| hdr   | str   | Text to be displayed for the sheet header.                                                        |
+-------+-------+---------------------------------------------------------------------------------------------------+
| width | int   | Column width                                                                                      |
+-------+-------+---------------------------------------------------------------------------------------------------+
"""
_sheet_desc_d = dict(
    director=(
        dict(width=2),
        dict(desc='port', i=0, hdr='Port', width=5),
        dict(desc='did', i=0, hdr='DID (Hex)', width=6),
        dict(desc='port_addr', i=0, hdr='Port Addr (Hex)', width=6),
        dict(desc='index', i=0, hdr='Index', width=6),
        dict(desc='link_addr', i=0, hdr='Link Addr', width=7),
        dict(desc='fid', i=0, hdr='FID', width=5),
        dict(desc='cu', i=0, hdr='Attached Device', width=32),
        dict(desc='port_name', i=0, hdr='Port Name', width=32),
        dict(desc='vc_low', i=0, hdr='Low Qos VC', width=5),
        dict(desc='vc_med', i=0, hdr='Med Qos VC', width=5),
        dict(desc='vc_high', i=0, hdr='High Qos VC', width=5),
        dict(width=2),
        dict(desc='port', i=1, hdr='Port', width=5),
        dict(desc='did', i=1, hdr='DID (Hex)', width=6),
        dict(desc='port_addr', i=1, hdr='Port Addr (Hex)', width=6),
        dict(desc='index', i=1, hdr='Index', width=6),
        dict(desc='link_addr', i=1, hdr='Link Addr', width=7),
        dict(desc='fid', i=1, hdr='FID', width=5),
        dict(desc='cu', i=1, hdr='Attached Device', width=32),
        dict(desc='port_name', i=1, hdr='Port Name', width=32),
        dict(desc='vc_low', i=1, hdr='Low Qos VC', width=5),
        dict(desc='vc_med', i=1, hdr='Med Qos VC', width=5),
        dict(desc='vc_high', i=1, hdr='High Qos VC', width=5),
    ),
    fixed=(
        dict(width=2),
        dict(desc='port', i=0, hdr='Port', width=5),
        dict(desc='did', i=0, hdr='DID (Hex)', width=6),
        dict(desc='port_addr', i=0, hdr='Port Addr (Hex)', width=6),
        dict(desc='index', i=0, hdr='Index', width=6),
        dict(desc='link_addr', i=0, hdr='Link Addr', width=7),
        dict(desc='fid', i=0, hdr='FID', width=5),
        dict(desc='cu', i=0, hdr='Attached Device', width=32),
        dict(desc='port_name', i=0, hdr='Port Name', width=32),
        dict(desc='vc_low', i=0, hdr='Low Qos VC', width=5),
        dict(desc='vc_med', i=0, hdr='Med Qos VC', width=5),
        dict(desc='vc_high', i=0, hdr='High Qos VC', width=5),
    )
)


def _iocp_summary(iocp_obj_l, switch_map_d):
    """Returns a dictionary as follows: key=DID, value is a dictionary whose key is the port address and value is:

    Returns a dictionary of switches by domain ID. The key is the DID and value is a dictionary as follows:

    +-----------+-------+-------------------------------------------------------------------------------------------+
    | key       | Type  | Value                                                                                     |
    +===========+=======+===========================================================================================+
    | chpid_d   | dict  | Dictionary of CHPIDs. The key is the equivalent RNID tag for the CHPID.                   |
    |           |       | Each dictionary is as follows:                                                            |
    |           |       | Key       Type    Value                                                                   |
    |           |       | iocp      str     Name of the CPC (server) for this CHPID                                 |
    |           |       | lpars_l   list    LPARS using this CHPID                                                  |
    |           |       | link_l    list    link addresses (keys for link_d)                                        |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | link_d    | dict  | Essentially the reverse of chpid_d. The key is the 2-byte link address. The value is a    |
    |           |       | a dictionaries as follows:                                                                |
    |           |       | Key       Type    Value                                                                   |
    |           |       | chpid_l   list    CHPIDs. Text is CPC: CSS CHPID                                          |
    |           |       | cu_type_l list    Control unit types. The only time more than one is valid is if CUP,     |
    |           |       |                   type 2032, is also defined.                                             |
    |           |       | cu_num_l  list    Control unit numbers as str.                                            |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | error_d   | dict  | Error messages.                                                                           |
    +-----------+-------+-------------------------------------------------------------------------------------------+

    :param iocp_obj_l: List of brcddb.classes.iocp.IOCPObj
    :type iocp_obj_l: list
    :param switch_map_d: Switch map. Key is the Switch ID and value is the DID
    :type switch_map_d: dict
    :return: Dictionary as defined above
    :rtype: dict
    """
    bad_id_l, r_switch_d = list(), dict()  # r_switch_d is the return dictionary

    for iocp_obj in iocp_obj_l:
        iocp = str(iocp_obj.r_obj_key())
        brcdapi_log.log('Processing: ' + iocp, True)

        # Spin through all the CHPIDs in the IOCP and determine the switch from the link addresses in the paths
        for chpid_obj in iocp_obj.r_path_objects():

            # Get or create the switch dictionary
            switch_did = chpid_obj.r_switch_id()
            if switch_did in switch_map_d:
                switch_did = switch_map_d[switch_did]
            try:
                x = int(switch_did, 16)
            except ValueError:
                x = 0
            if x < 1 or x > 239:
                # switch_did = '00'
                buf = 'Switch ID ' + switch_did + ' in ' + iocp
                bad_id_l.append(buf)
            switch_d = r_switch_d.get(switch_did)
            if switch_d is None:
                switch_d = dict(chpid_d=dict(), link_d=dict(), error_d=dict())
                r_switch_d.update({switch_did: switch_d})
            if switch_did == '00':
                error_key = len(switch_d['error_d'].keys())
                buf = 'Invalid switch ID, ' + chpid_obj.r_switch_id() + ' in ' + iocp
                switch_d['error_d'].update({error_key: dict(area='Switch', type='ID', desc=buf)})

            # Add a CHPID dictionary to the switch dictionary
            chpid_d = dict(iocp=iocp, lpars_l=chpid_obj.r_lpars(), link_l=list())
            switch_d['chpid_d'].update({chpid_obj.r_obj_key(): chpid_d})

            # Figure out all the DIDs from the link addresses and get a switch dictionary for switch_d
            for path_addr in chpid_obj.r_link_addresses():
                path_d = chpid_obj.r_link_addr(path_addr)

                # Get the switch dictionary for the DID in the link address
                link_addr = path_addr if len(path_addr) == 4 else switch_did + path_addr
                chpid_d['link_l'].append(link_addr)
                did = link_addr[0:2]
                switch_d = r_switch_d.get(did)
                if switch_d is None:
                    switch_d = dict(chpid_d=dict(), link_d=dict(), error_d=dict())
                    r_switch_d.update({did: switch_d})

                # Get the link & control unit dictionaries
                link_d = switch_d['link_d'].get(link_addr)
                if link_d is None:
                    link_d = dict(chpid_l=list(), cu_type_l=list(), cu_num_l=list())
                    switch_d['link_d'].update({link_addr: link_d})
                link_d['chpid_l'].append(iocp_obj.r_obj_key() + ': ' + brcddb_iocp.tag_to_text(chpid_obj.r_obj_key()))
                link_d['cu_num_l'].extend([str(k) for k in path_d.keys()])
                link_d['cu_type_l'].extend(path_d.values())

    # Let the user know if there are any Switch IDs that cannot be converted to DIDs
    if len(bad_id_l) > 0:
        buf = 'Cannot convert the following Switch IDs, "SWITCH=" statement in the CNTLUNIT macro, to a valid Domain '\
              'ID (DID). Use a switch map, -map, to convert Switch IDs to Domain IDs for the following:'
        brcdapi_log.log(['', buf] + ['  ' + b for b in gen_util.remove_duplicates(bad_id_l)] + [''], True)

    # Remove all the duplicates. Note that multiple CHPIDs can have a path to the same link address
    for switch_d in r_switch_d.values():
        for k0, v0 in switch_d.items():
            if isinstance(v0, list):
                switch_d[k0] = gen_util.remove_duplicates(v0)
            elif isinstance(v0, dict):
                for k1, v1 in v0.items():
                    if isinstance(v1, list):
                        v0[k1] = gen_util.remove_duplicates(v1)

    return r_switch_d


def _about_sheet(wb, sheet_index):
    """Inserts a worksheet into the workbook containing basic information and instructions

    :param wb: openpyxl Workbook
    :type wb: openpyxl class object
    :param sheet_index: Index as to where to start inserting worksheets
    :type sheet_index: int
    """
    global _hdr1_font, _hdr2_font, _bold_font, _std_font, _alignment

    # Add an "about" sheet
    sheet = wb.create_sheet(index=sheet_index, title='About')
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.column_dimensions['A'].width = 80

    row = col = 1
    for d in _about_sheet_l:
        if d is not None:
            excel_util.cell_update(sheet, row, col, d['t'], font=d['f'], align=_alignment)
        row += 1


# Case statements for _summary_sheet()
def _summary_chpid(k, d):
    return k, d['iocp'] + ': ' + brcddb_iocp.tag_to_text(k), ', '.join(d['link_l'])


def _summary_link(k, d):
    return (
        k,
        ', '.join(gen_util.remove_duplicates(d['cu_type_l'])),
        ', '.join(gen_util.remove_duplicates(d['cu_num_l']))
    )


def _summary_errors(k, d):
    return d['area'], d['type'], d['desc']


_summary_case = dict(
    chpid_d=_summary_chpid,
    link_d=_summary_link,
    error_d=_summary_errors,
)


def _summary_sheet(wb, sheet_index, did, did_d):
    """Inserts a summary worksheet into the workbook

    :param wb: openpyxl Workbook
    :type wb: openpyxl class object
    :param sheet_index: Index as to where to start inserting worksheets
    :type sheet_index: int
    :param did: Domain ID in hex. No leading '0x'
    :type did: str
    :param did_d: Definition for this switch (individual dictionary returned from _iocp_summary()
    :type did_d: dict
    """
    global _hdr1_font, _hdr2_font, _bold_font, _std_font, _alignment

    # Add a summary sheet
    sheet = wb.create_sheet(index=sheet_index, title='Summary')
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    col = 1
    for x in _summary_sheet_d['col']:
        sheet.column_dimensions[xl.get_column_letter(col)].width = x
        col += 1
    end_col = col - 1
    row = col = 1

    # Add the Domain ID as the title
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    excel_util.cell_update(sheet, row, col, 'Domain ID: ' + str(did), font=_hdr2_font, align=_alignment)
    row += 1

    # Fill out the CHPIDs, control units, and bad links
    for k in ('chpid_d', 'link_d', 'error_d'):
        row += 1

        # The Title
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
        excel_util.cell_update(sheet, row, 1, _summary_sheet_d[k]['t'].replace('$c', str(len(did_d[k].keys()))),
                               font=_hdr2_font, align=_alignment)

        # The column headers
        row, col = row + 1, 1
        for buf in _summary_sheet_d[k]['h']:
            excel_util.cell_update(sheet, row, col, buf, font=_bold_font, align=_alignment)
            col += 1

        # The content
        row += 1
        for item_k, item_d in did_d[k].items():
            col = 1
            for buf in _summary_case[k](item_k, item_d):
                excel_util.cell_update(sheet, row, col, buf, font=_std_font, align=_alignment)
                col += 1
            row += 1


def _switch_config(wb, sheet_index, sheet_name, did, sheet_l):
    """Create the workbook and add the Instructions, Chassis, and Switch_x worksheets
    
    :param wb: openpyxl Workbook
    :type wb: openpyxl class object
    :param sheet_index: Index as to where this sheet should be placed in the Workbook
    :type sheet_index: int
    :param sheet_name: Name of worksheet
    :type sheet_name: str
    :param did: DID to initialize in hex
    :param did: str
    :param sheet_l: Sheet list returned from excel_util.read_workbook
    :type sheet_l: list
    """
    global _switch_config_hdr, _switch_config_d, _hdr1_font, _std_font, _border, _alignment, _dv_yn, _dv_fid, _dv_did
    global _dv_switch_type, _dv_dup_wwn, _dv_fab_principal, _dv_routing_policy, _dv_port_name

    # Add the worksheet and data validations
    sheet = wb.create_sheet(index=sheet_index, title=sheet_name)
    sheet.page_setup.paperSize = sheet.PAPERSIZE_LETTER
    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
    for dv in (_dv_yn, _dv_fid, _dv_did, _dv_switch_type, _dv_dup_wwn, _dv_routing_policy, _dv_port_name,
               _dv_fab_principal):
        sheet.add_data_validation(dv)

    # Add the column headers
    row = col = 1
    for d in _switch_config_hdr:
        sheet.column_dimensions[xl.get_column_letter(col)].width = d['c']
        excel_util.cell_update(sheet, row, col, d['t'], font=_hdr1_font, align=d['align'], border=_border)
        col += 1
    row, col = row + 1, 1

    # Find Switch_x and add the body
    for sheet_d in sheet_l:
        if str(sheet_d.get('sheet')) == 'Switch_x':
            for row_l in sheet_d['al'][1:]:
                try:
                    d = dict() if _switch_config_d.get(row_l[0]) is None else _switch_config_d[row_l[0]]
                    excel_util.cell_update(sheet, row, col, row_l[0], font=_std_font, align=_alignment, border=_border)
                    col += 1
                    v = '0x' + did if row_l[0] == 'Domain ID (DID)' else d['p'] if 'p' in d else row_l[1]
                    excel_util.cell_update(sheet, row, col, v, font=_std_font, align=_alignment, border=_border)
                    dv = d.get('dv')
                    if dv is not None:
                        dv.add(xl.get_column_letter(col) + str(row))
                    col += 1
                    excel_util.cell_update(sheet, row, col, row_l[2], font=_std_font, align=_alignment, border=_border)
                except IndexError:
                    brcdapi_log.log('Malformed sheet in template: Switch_x', echo=True)
                    break
                row, col = row + 1, 1
            return

    brcdapi_log.log('Could not find Switch_x in template', echo=True)


def _port_sheets(wb, sheet_index, switch_type, did, did_d):
    """Creates a worksheet and fills in the content for each port card ("Slot x").

    :param wb: openpyxl Workbook
    :type wb: openpyxl class object
    :param sheet_index: Index as to where to start inserting worksheets
    :type sheet_index: int
    :param switch_type: The switch type. See _config_d
    :type switch_type: str
    :param did: Domain ID in hex. No leading '0x'
    :type did: str
    :rtype: None
    """
    global _config_d, _generic_blade_type_d, _row_val_d, _com_switch_type_d, _special_d, _fill_d
    global _center_align, _white_bold_font, _border, _config_slot, _std_font, _fill_asic_d

    # Build a cross-reference dictionary using the cell as a key and the value the content
    link_d = dict()
    for link_addr, cu_d in did_d['link_d'].items():
        cu_type = ', '.join(gen_util.remove_duplicates(cu_d['cu_type_l']))
        if cu_type != '2032':  # The CUP port doesn't need a physical port
            # Keep track of where all the control units go and make sure the switch can accommodate the link address
            buf = 'Type: ' + cu_type + '\n' + 'CUNUM: ' + ', '.join(cu_d['cu_num_l']) + '\n' + \
                  '\n'.join(cu_d['chpid_l'])
            link_d.update({link_addr: buf})

    # Read the template
    local_did = '00'
    common_switch_type = _com_switch_type_d[switch_type]
    error_l, sheet_l = excel_util.read_workbook(_config_d[switch_type], dm=3, order='col', skip_sheets=_skip_sheets)
    if len(error_l) > 0:
        brcdapi_log.exception(error_l, echo=True)
        return
    for sheet_d in [d for d in sheet_l if d['sheet'][0: min(len(d['sheet']), len('Slot '))] == 'Slot ']:

        # Set up the local data structures
        slot = int(sheet_d['sheet'].split(' ')[1])
        generic_blade_type = _generic_blade_type_d[common_switch_type][slot]
        switch_d = _row_val_d[generic_blade_type]
        row_l = sheet_d['al']
        start_col = switch_d['col']['start']
        end_col = switch_d['col']['end']
        col_keys_l = [int(i) for i in switch_d.keys() if str(i).isnumeric() and not switch_d[i].get('skip', False)]

        # Create a worksheet, set the column widths, and add the title
        sheet = wb.create_sheet(index=sheet_index, title='Slot ' + str(slot))
        sheet_index += 1
        for col in [i for i in switch_d.keys() if str(i).isnumeric()]:
            sheet.column_dimensions[xl.get_column_letter(col+1)].width = switch_d[col]['width']
        excel_util.cell_update(sheet, 1, 2, row_l[0][1], align=_center_align, font=_white_bold_font, border=_border,
                               fill=_config_slot)
        sheet.merge_cells(start_row=1, start_column=start_col+1, end_row=1, end_column=end_col)

        # Add the headers
        for col in col_keys_l:
            excel_util.cell_update(sheet, 2, col+1, row_l[1][col], align=_center_align, font=_std_font, border=_border)

        # Add the content
        for row in reversed(range(switch_d['row']['start'], switch_d['row']['end'])):
            cu, col_l = None, row_l[row]
            for col in col_keys_l:
                if switch_d[col].get('port_addr', False):
                    cu = link_d.get(local_did + gen_util.pad_string(str(col_l[col]), pad_len=2, pad_char='0'))
                try:
                    val = _special_d[common_switch_type][generic_blade_type][slot][row][col]
                except KeyError:
                    try:
                        val = switch_d[col]['val']
                    except KeyError:
                        val = cu if switch_d[col].get('ad', False) else col_l[col]
                if isinstance(val, str):  # +1 for $r and +2 for $nr because these are Excel rows, not zero based
                    if switch_d[col].get('did', False):
                        if val == '$did':
                            local_did = did
                        elif val == '$tv':
                            local_did = gen_util.pad_string(str(col_l[col]), 2, '0')
                    val = col_l[col] if val == '$tv' \
                        else did.lstrip('0') if val == '$did' \
                        else val.replace('$nr', str(row+2)).replace('$r', str(row+1))
                excel_util.cell_update(sheet, row+1, col+1, val, align=switch_d[col].get('align'), font=_std_font,
                                       border=_border, fill=_fill_asic_d[_fill_d[generic_blade_type][row][col]])

    return


def _write_report(summary_d, prefix, switch_type, sheet_l):
    """Generate a dictionary summarizing the switch requirements based on the IOCP objects

    :param summary_d: Value returned form _iocp_summary()
    :type summary_d: dict
    :param prefix: Prefix for name of Excel file for the report. It is appended with '_xx' where xx is the DID
    :type prefix: str
    :param switch_type: Type of switch. See _switch_types
    :type switch_type: str
    :param sheet_l: Sheets returned from excel_util.read_workbook
    :type sheet_l: list
    :return: Error messages. Empty list if no errors encountered
    :rtype: list
    """
    global _copy_sheets_d, _hdr1_font, _std_font, _border, _alignment

    # Set up the workbook and add a switch configuration worksheet template for each switch by DID
    rl = list()
    for did, did_d in summary_d.items():
        wb_name = prefix + '_' + did + '.xlsx'
        brcdapi_log.log('Creating Workbook: ' + wb_name, True)
        wb = xl_wb.Workbook()
        sheet_index = 0
        for sheet_name, d in _copy_sheets_d.items():
            rl.extend(excel_util.copy_worksheet(wb,
                                                sheet_index,
                                                sheet_name,
                                                sheet_l,
                                                col_width_l=d.get('width_l'),
                                                font=d.get('font'),
                                                fill=d.get('fill'),
                                                border=d.get('border'),
                                                align=d.get('align')))
        sheet_index += 1
        _switch_config(wb, sheet_index, 'Switch_x', did, sheet_l)
        sheet_index += 1
        _port_sheets(wb, sheet_index, switch_type, did, did_d)
        sheet_index = 1
        _summary_sheet(wb, sheet_index, did, did_d)
        sheet_index = 0
        _about_sheet(wb, sheet_index)
        try:
            wb.save(wb_name)
        except FileNotFoundError:
            rl.append('The folder in ' + wb_name + ' does not exist.')
        except PermissionError:
            rl.append('Write permission for ' + wb_name + ' denied. This is typically due to the file being opened.')

    return rl


def psuedo_main(proj_obj, prefix, switch_type, iocp, switch_id_map, sheet_l, debug_flag):
    """Basically the main(). Did it this way so that it can easily be used as a standalone module or called from another

    :param proj_obj: Project object. If None, an error was encountered.
    :type proj_obj: None, brcddb.classes.project.ProjectObj
    :param prefix: Switch workbook prefix
    :type prefix: str
    :param switch_type: Switch type
    :type switch_type: str
    :param iocp: Folder containing IOCP files
    :type iocp: str
    :param switch_id_map: Dictionary of switch IDs to domain IDs
    :type switch_id_map: dict
    :param sheet_l: Workbook template sheets read from excel_util.read_workbook()
    :type sheet_l: list
    :param debug_flag: If True, write IOCP files with just CNTLUNIT and PATH=
    :return: Exit code. See exist codes in brcddb.brcddb_common
    :rtype: int
    """
    global __version__

    # Parse the IOCP files
    file_l = brcdapi_file.read_directory(iocp)
    if len(file_l) == 0:
        brcdapi_log.log(['', 'No IOCP files found in ' + iocp, ''], True)
        return brcddb_common.EXIT_STATUS_INPUT_ERROR

    if debug_flag:
        for file in file_l:
            brcdapi_log.log('Parsing: ' + file, True)
            iocp_l = brcdapi_file.read_file(iocp + '/' + file, False, False)
            chpid_l, cntlunit_l = brcddb_iocp.condition_iocp(iocp_l)
            try:
                brcdapi_file.write_file(
                    prefix + file.split('.')[0] + '.txt',
                    [buf.strip() for buf in cntlunit_l if 'LINK=' in buf],
                )
            except (FileExistsError, FileNotFoundError, PermissionError):
                brcdapi_log.log('Error writing file: ' + file, echo=True)
            return brcddb_common.EXIT_STATUS_OK

    for file in file_l:
        brcdapi_log.log('Parsing: ' + file, True)
        brcddb_iocp.parse_iocp(proj_obj, iocp + '/' + file)

    # Generate the report
    ml = _write_report(_iocp_summary(proj_obj.r_iocp_objects(), switch_id_map), prefix, switch_type, sheet_l)

    if len(ml) > 0:
        brcdapi_log.log(ml, True)
        return brcddb_common.EXIT_STATUS_ERROR
    return brcddb_common.EXIT_STATUS_OK


def _get_input():
    """Gets the shell invocation parameters, validates input, opens log file, gets project object, gets switch map

    :return: Exit code. See exist codes in brcddb.brcddb_common
    :rtype: int
    """
    global __version__, _input_d

    proj_obj, error_l, switch_map_d, sheets_l, el = None, list(), dict(), list(), list()

    # Get command line input
    buf = 'Parses IOCP files and generates planning workbooks. WARNING: For openpyxl to read the workbooks generated '\
          'by this script, which uses the same openpyxl library, the files must be opened and saved in Excel. There '\
          'is no need to make any changes. This affects switch_confi.py and switch_config_cli.py.'
    args_d = gen_util.get_input(buf, _input_d)

    # Set up logging
    brcdapi_log.open_log(
        folder=args_d['log'],
        suppress=args_d['sup'],
        no_log=args_d['nl'],
        version_d=brcdapi_util.get_import_modules()
    )

    # User feedback
    ml = [
        os.path.basename(__file__) + ', ' + __version__,
        'Workbook prefix, -p:        ' + args_d['p'],
        'Switch type, -t:            ' + args_d['t'],
        'IOCP folder, -iocp:         ' + args_d['iocp'],
        'Switch ID to DID map, -map: ' + str(args_d['map']),
        'Template file, -t:          ' + str(_config_d.get(args_d['t'])),
        'Debug, -d:                  ' + str(args_d['d']),
        'Log, -log:                  ' + str(args_d['log']),
        'No log, -nl:                ' + str(args_d['nl']),
        'Suppress, -sup:             ' + str(args_d['sup']),
        '',
    ]
    brcdapi_log.log(ml, echo=True)

    # Parse the switch ID to DID map
    if isinstance(args_d['map'], str):
        for buf in args_d['map'].split(';'):
            temp_l = buf.split(',')
            if len(temp_l) != 2:
                error_l.append(buf + ' is not valid in ' + args_d['map'])
            else:
                try:
                    x = int(temp_l[1], 16)
                    if x < 1 or x > 239:
                        raise InvalidDID
                except (InvalidDID, ValueError):
                    error_l.append('Domain ID ' + temp_l[1] + ' is not valid in ' + args_d['map'])
            if len(error_l) == 0:
                switch_map_d.update({temp_l[0]: '0' + temp_l[1] if len(temp_l[1]) == 1 else temp_l[1]})

    # Read in template
    temp_file = brcdapi_file.full_file_name(_config_d[args_d['t']], '.xlsx')
    try:
        el, sheets_l = excel_util.read_workbook(temp_file, dm=3, order='col', skip_sheets=_skip_sheets)
        error_l.extend(el)
    except FileExistsError:
        error_l.append('A folder in template file does not exist: ' + temp_file)
    except FileNotFoundError:
        error_l.append('Template file not found: ' + temp_file)
    except PermissionError:
        error_l.append('You do not have access rights to read ' + temp_file + '.')

    # Get a project object and run the script if there were no errors
    if len(error_l) == 0:
        proj_obj = brcddb_project.new('parse_iocp.py', datetime.datetime.now().strftime('%Y_%m_%d_%H_%M_%S'))
        proj_obj.s_python_version(sys.version)
        proj_obj.s_description('Parsed IOCPs')
        return psuedo_main(proj_obj, args_d['p'], args_d['t'], args_d['iocp'], switch_map_d, sheets_l, args_d['d'])

    brcdapi_log.log(error_l, echo=True)
    return brcddb_common.EXIT_STATUS_INPUT_ERROR


##################################################################
#
#                    Main Entry Point
#
###################################################################

# Read in the project file from which the report is to be created and convert to a project object
# Create project

if _DOC_STRING:
    print('_DOC_STRING is True. No processing')
    exit(brcddb_common.EXIT_STATUS_OK)
else:
    _ec = _get_input()
    brcdapi_log.close_log('Processing complete with ending status: ' + str(_ec), echo=True)
    exit(_ec)
