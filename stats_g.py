#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Copyright 2023, 2024, 2025 Consoli Solutions, LLC.  All rights reserved.

**License**

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

**Description**

Add statistics to Excel Workbook

Reads in the output of stats_c (which collects port statistics) and creates an Excel Workbook for each port.

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 03 Apr 2024   | Added version numbers of imported libraries.                                          |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 29 Oct 2024   | Fixed call to cell_match_val().                                                       |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 06 Dec 2024   | Fixed spelling mistake in message.                                                    |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.5     | 25 Aug 2025   | Use brcddb.util.util.get_import_modules to dynamically determined imported libraries. |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2024, 2025 Consoli Solutions, LLC'
__date__ = '25 Aug 2025'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack_consoli@yahoo.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.5'

import sys
import os
import datetime
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import openpyxl.utils.cell as xl
import brcdapi.log as brcdapi_log
import brcdapi.gen_util as gen_util
import brcdapi.util as brcdapi_util
import brcdapi.excel_util as excel_util
import brcdapi.file as brcdapi_file
import brcddb.brcddb_project as brcddb_project
import brcddb.util.util as brcddb_util
import brcddb.brcddb_common as brcddb_common
import brcddb.report.utils as report_utils
import brcddb.app_data.report_tables as rt
import brcddb.report.port as report_port
import brcddb.util.copy as brcddb_copy
import brcddb.brcddb_port as brcddb_port
import brcddb.brcddb_fabric as brcddb_fabric
import brcddb.util.search as brcddb_search
import brcddb.report.graph as report_graph

_DOC_STRING = False  # Should always be False. Prohibits any code execution. Only useful for building documentation
# _STAND_ALONE: True: Executes as a standalone module taking input from the command line. False: Does not automatically
# execute. This is useful when importing this module into another module that calls psuedo_main().
_STAND_ALONE = True  # See note above

_buf = 'Optional. Specifies the graph type. Valid chart types are: ' + ', '.join(report_graph.chart_types) + '. '
_buf += 'Default: line.'
_input_d = dict(
    r=dict(h='Required. Report name. ".xlsx" is automatically appended.'),
    i=dict(h='Required. Name of data input file. This must be the output file, -o, from stats_c.py. ".json" is '
             'automatically appended'),
    gp=dict(r=False,
            h='Optional. Creates a worksheet with a graph of one or more statistical counters for specific ports. '
              'Parameters that follow are the port number followed by any of the statistical parameters in '
              'brocade-interface/fibrechannel-statistics. Only the final leaf should be used. All parameters must be '
              'separated by a comma. Separate ports with a semi-colon. For example, to graph the Tx and Rx frames for '
              'port 3/14: "-gp 3/14,in-frames,out-frames"'),
    gs=dict(r=False,
            h='Optional. Name of the worksheet to create with a graph of one or more statistical counters. This is '
              'essentially the reverse of the -gp option. By default, all ports are plotted for the specified '
              'statistic. You can filter ports by specifying "top-x" or "avg-x", where "x" is the number of ports to '
              'graph. "top-x" looks for the ports with the top highest peaks while "avg-x" looks for the highest total '
              'count for the specified statistic. For example, to plot in-frame and out-frames for all ports: '
              '"-gs in-frames;out-frames". For the same graph, but limited to 5 ports with the highest average '
              '(which is the same as the highest total count for all samples): "-gs '
              'in-frames,avg-5;out-frames,avg-5".'),
    gt=dict(r=False, h=_buf),
)
_input_d.update(gen_util.parseargs_log_d.copy())

_invalid_parm_ref = 'Invalid display parameter: '
_invalid_port_ref = 'Invalid port reference: '
_sheet_map = dict()  # key: port number, value: openpyxl sheet

_port_stats = (
    'fibrechannel/average-transmit-frame-size',
    'fibrechannel/average-receive-frame-size',
    'fibrechannel/average-transmit-buffer-usage',
    'fibrechannel/average-receive-buffer-usage'
    # 'fibrechannel/current-buffer-usage',
    # 'fibrechannel/recommended-buffers',
    # 'fibrechannel/chip-buffers-available',
) + rt.Port.port_stats1_tbl


# Case methods for _get_ports(). See _port_match
def _e_ports(switch_obj):
    return [p.r_obj_key() for p in brcddb_search.match_test(switch_obj.r_port_objects(), brcddb_search.e_ports)]


def _f_ports(switch_obj):
    return [p.r_obj_key() for p in brcddb_search.match_test(switch_obj.r_port_objects(), brcddb_search.f_ports)]


def _all(switch_obj):
    return switch_obj.r_port_keys()


def _ports(proj_obj, port):
    global _invalid_port_ref

    r = list()
    switch_obj = proj_obj.r_switch_obj(proj_obj.r_get('base_switch_wwn'))
    s, p = gen_util.slot_port(port)
    if s is None or p is None:
        if gen_util.is_valid_zone_name(port):
            fab_obj = switch_obj.r_fabric_obj()
            ml = list()  # Fill this with login WWNs
            # Remember that a zone name can't be the same as an alias name so one of these loops will do nothing
            for obj in fab_obj.r_zones_for_wwn(port):  # Zone list for alias. Yes, port can be a WWN or an alias
                l = obj.r_members() + obj.r_pmembers()
                for mem in l:
                    if gen_util.is_wwn(mem):
                        ml.append(mem)
                    else:
                        obj = fab_obj.r_alias_obj(mem)
                        if obj is not None:
                            ml.extend(obj.r_members())
            obj = fab_obj.r_alias_obj(port)  # Aliases
            if obj is not None:
                ml.extend(obj.r_members())

            # Now get all the port numbers associated with the WWNs in ml
            switch_wwn = switch_obj.r_obj_key()
            for login_obj in [fab_obj.r_login_obj(mem) for mem in ml if fab_obj.r_login_obj(mem) is not None]:
                port_obj = login_obj.r_port_obj()
                if port_obj is not None and port_obj.r_switch_key() == switch_wwn:
                    # When written, login objects were only returned on the switch where the login occurred so
                    # switch_wwn has to match port_obj.r_switch_key() so this test is future proofing
                    r.append(port_obj.r_obj_key())
        else:
            brcdapi_log.log(_invalid_port_ref + port, echo=True)
    else:
        if switch_obj.r_port_obj(port) is None:
            brcdapi_log.log('Port ' + port + ' not found', echo=True)
        else:
            return list(port)
    return list()  # If we get here, something went wrong


_port_match = {
    'E-PORTS': _e_ports,
    'F-PORTS': _f_ports,
    'ALL': _all,
}


def _get_ports(switch_obj, ports):
    """Returns a list of port objects matching the user input.

    :param switch_obj: Project object
    :type switch_obj: brcddb.classes.ProjectObj
    :param ports: Port or list of port types. See -p parameter in parse_args()
    :type ports: str
    :return: List of port objects (brcddb.classes.PortObj)
    :rtype: list
    """
    global _port_match

    r = list()
    for port in gen_util.convert_to_list(ports.split(',')):
        if port in _port_match:
            r.extend(_port_match.get(port)(switch_obj))
        else:
            r.extend(_ports(switch_obj, port))

    r.extend(r)
    return gen_util.remove_duplicates(r)


def _get_parameters(parms):
    """Returns a list of Excel column letters matching the list of parms

    :param parms: Parameters to plot as passed in from the command line with the -p option
    :type parms: None, str, list
    :return: List of columns matching the parameters
    :rtype: list
    """
    global _invalid_parm_ref

    # Build a header to KPI reference table
    r_map = dict()
    for k, d in rt.Port.port_display_tbl.items():
        if 'fibrechannel-statistics/' in k:
            v = d.get('d')
            if v is not None:
                r_map.update({v: k})

    # Now figure out what to return
    r = list()
    for p in parms.split(','):
        v = r_map.get(p)
        if v is None:
            brcdapi_log.log(_invalid_parm_ref + p, echo=True)
        else:
            r.append(v)
    return r


def _add_ports(wb, tc_page, t_content, start_i, switch_obj):
    """Add the individual port pages to the workbook

    :param wb: Excel workbook object
    :type wb: Workbook object
    :param tc_page: Name of table of contents page
    :type tc_page: str
    :param t_content: Table of contents
    :type t_content: list
    :param start_i: Starting index (where first port goes)
    :param switch_obj: Base switch object
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :rtype: None
    """
    global _sheet_map, _port_stats

    sheet_index = start_i
    proj_obj = switch_obj.r_project_obj()
    switch_obj_l = [proj_obj.r_switch_obj(wwn) for wwn in proj_obj.r_get('switch_list')]
    for port_obj in brcddb_util.sort_ports(switch_obj.r_port_objects()):

        # Create the port page
        port_num = port_obj.r_obj_key()
        sname = port_num.replace('/', '_')
        brcdapi_log.log('Processing port: ' + port_num, echo=True)
        port_list = [obj.r_port_obj(port_num) for obj in switch_obj_l]
        sheet = report_port.port_page(wb, '#' + tc_page + '!A1', sname, sheet_index, 'Port: ' + port_num, port_list,
                                      _port_stats, rt.Port.port_display_tbl, False)
        _sheet_map.update({port_num: sheet})

        # Add the port page to the table of contents
        t_content.append(dict(new_row=False, font='link', align='wrap', hyper='#' + sname + '!A1', disp=port_num))
        t_content.append(dict(new_row=False, font='std', align='wrap', disp=port_obj.r_port_name()))
        t_content.append(dict(new_row=False, font='std', align='wrap', disp=port_obj.c_login_type()))
        v = port_obj.r_get('fibrechannel/operational-status')
        try:
            buf = brcddb_common.port_conversion_tbl['fibrechannel/operational-status'][v]
        except KeyError:
            buf = 'Unknown'
        t_content.append(dict(new_row=False, font='std', align='wrap', disp=buf))
        t_content.append(dict(font='std', align='wrap', disp=brcddb_port.port_best_desc(port_obj)))
        sheet_index += 1


def _add_graphs(wb, tc_page, t_content, start_i, base_switch_obj, graph_list):
    """Add the individual port pages to the workbook

    Note: If there is a way to add multiple lines to a graph that aren't in neighboring columns, I haven't figured out
    how. So what we do here is add a worksheet that is a copy of the data from the different ports and or statistics
    and add it to the end. If there was just one item to plot, I wouldn't have to add another sheet but this algorithm
    is simple. It just creates a worksheet and adds columns as necessary, even if it's just one.

    :param wb: Excel workbook object
    :type wb: Workbook object
    :param tc_page: Name of table of contents page
    :type tc_page: str
    :param t_content: Table of contents
    :type t_content: list
    :param start_i: Starting index (where first port goes)
    :param base_switch_obj: Base switch object
    :type base_switch_obj: brcddb.classes.switch.SwitchObj
    :param graph_list: List of graph dictionaries, see graph in _write_report() for details
    :type graph_list: list, tuple
    :return: List of warnings or errors.
    :rtype: list
    """
    global _sheet_map

    ml = list()
    proj_obj = base_switch_obj.r_project_obj()
    switch_obj_l = [proj_obj.r_switch_obj(wwn) for wwn in proj_obj.r_get('switch_list', list())]
    if len(switch_obj_l) < 2:
        brcdapi_log.log('Nothing to graph. No data collected.', echo=True)
        return ml
    sheet_index = start_i
    last_disp = {'font': 'std', 'align': 'wrap'}
    std_disp = last_disp.copy()
    std_disp.update({'new_row': False})

    graph_num = 0
    for graph_obj in graph_list:

        # Create the graph data page and figure out the common graphing request
        sname = 'graph_' + str(graph_num)
        brcdapi_log.log('Processing graph for: ' + sname, echo=True)
        data_sheet = wb.create_sheet(title=sname + '_data')
        data_sheet.page_setup.paperSize = data_sheet.PAPERSIZE_LETTER
        data_sheet.page_setup.orientation = data_sheet.ORIENTATION_LANDSCAPE
        port, stat = graph_obj.get('port'), graph_obj.get('stat')
        y_name = 'Programming error. Neither port or stat specified.'
        title, last_time = '', ''

        if port is not None:

            # Set up the title and chart references
            switch_obj = switch_obj_l[0]
            port_obj = switch_obj.r_port_obj(port)
            if port_obj is None:
                ml.append('Statistics for port ' + port + ' were not collected. Skipping.')
                continue
            col_ref = port
            port_page = '=' + port.replace('/', '_') + '!'

            # Find all the time stamps, reference sheets, and column references for the data reference sheet
            rd = dict()
            sheet = _sheet_map[port]
            data_sheet['A1'] = 'Time'  # Column header for the time stamp
            col = 2
            for stat in gen_util.convert_to_list(graph_obj.get('parms')):
                try:
                    stat_ref = stat if rt.Port.port_display_tbl['fibrechannel-statistics/' + stat]['d'] is None else \
                        rt.Port.port_display_tbl['fibrechannel-statistics/' + stat]['d']
                except (ValueError, TypeError):
                    stat_ref = stat
                cell = excel_util.cell_match_val(sheet, stat_ref, None, 2, 1)
                if cell is None:
                    ml.append('Could not find statistical count ' + stat + ' for port ' + port + '. Skipping')
                    continue
                ref_col = column_index_from_string(coordinate_from_string(cell)[0])
                rd.update({stat: ref_col})
                data_sheet[xl.get_column_letter(col) + '1'] = port_page + cell
                col += 1
            max_col = len(rd.keys())

            # Add the time stamp
            x = port_obj.r_get('fibrechannel-statistics/time-generated')
            if x is None:
                ml.append('Invalid sample for port ' + port + '. Skipping.')
                break
            title = 'Statistics for port ' + port + 'beginning: ' + datetime.datetime.fromtimestamp(x).strftime(
                '%d %b %Y, %H:%M:%S')
            row = 2
            for port_obj in [obj.r_port_obj(port) for obj in switch_obj_l]:
                x = None if port_obj is None else port_obj.r_get('fibrechannel-statistics/time-generated')
                if x is None:
                    buf = 'Port ' + port + ' appears to have gone off line after ' + str(last_time)
                    buf += '. Switch: ' + port_obj.r_switch_obj().r_obj_key()
                    brcdapi_log.log(buf, echo=True)
                    break
                last_time = x
                data_sheet['A' + str(row)] = datetime.datetime.fromtimestamp(x).strftime('%H:%M:%S')
                col = 2
                for ref_col in rd.values():
                    data_sheet[xl.get_column_letter(col) + str(row)] = port_page + xl.get_column_letter(ref_col) +\
                                                                       str(row + 1)
                    col += 1
                row += 1

        elif stat is not None:

            # Figure out the title and graph Y axis title
            y_name = stat.split('/').pop()
            try:
                col_ref = y_name if rt.Port.port_display_tbl[stat]['d'] is None else rt.Port.port_display_tbl[stat]['d']
            except (ValueError, TypeError) as e:
                col_ref = y_name

            # Find all the time stamps, reference sheets, and columns
            tl, rl = list(), list()
            for port in gen_util.convert_to_list(graph_obj.get('parms')):
                port_obj = switch_obj_l[0].r_port_obj(port)
                if port_obj is None:
                    ml.append('Could not find port ' + port)
                    continue
                sheet = _sheet_map[port]
                cell = excel_util.cell_match_val(sheet, col_ref, None, 2, 1)
                if cell is None:
                    ml.append('Could not find column for port ' + port + ', statistic ' + stat)
                    continue
                rl.append(dict(sheet=sheet, port=port, name=port.replace('/', '_'),
                               col=column_index_from_string(coordinate_from_string(cell)[0])))
                if len(tl) == 0:
                    try:
                        x = switch_obj_l[0].r_port_obj(port).r_get('fibrechannel-statistics/time-generated')
                    except (ValueError, TypeError):
                        x = None
                    if x is None:
                        ml.append('Invalid sample for port ' + port + '. Skipping')
                        break
                    title = 'Statistics beginning: ' + datetime.datetime.fromtimestamp(x).strftime('%d %b %Y, %H:%M:%S')
                    for port_obj in [obj.r_port_obj(port) for obj in switch_obj_l]:
                        x = port_obj.r_get('fibrechannel-statistics/time-generated')
                        if x is None:
                            buf = 'Port ' + port + ' appears to have gone off line after ' + str(last_time)
                            buf += '. Switch: ' + port_obj.r_switch_obj().r_obj_key()
                            brcdapi_log.log(buf, echo=True)
                            break
                        last_time = x
                        tl.append(datetime.datetime.fromtimestamp(x).strftime('%H:%M:%S'))

            # Add all the data references
            max_col = len(rl)
            if max_col > 0:
                data_sheet['A1'] = 'Time'  # Column header for the time stamp
                for row in range(0, len(tl)):  # Fill in the time stamp.
                    data_sheet['A' + str(row + 2)] = tl[row]
                for col in range(0, len(rl)):  # Now add the port column headers
                    data_sheet[xl.get_column_letter(col+2) + '1'] = "'" + rl[col]['port']
                for row in range(0, len(switch_obj_l)):  # Now add all the statistical data for the ports
                    col = 2
                    for ref in rl:  # One for each port
                        data_sheet[xl.get_column_letter(col) + str(row + 2)] = \
                            '=' + ref['name'] + '!' + xl.get_column_letter(ref['col']) + str(row + 3)
                        col += 1

        else:
            brcdapi_log.exception(y_name, echo=True)
            continue

        # Create the Worksheet and add it to the table of contents
        max_row = len(switch_obj_l) + 1
        report_graph.graph(wb, '#' + tc_page + '!A1', sname, sheet_index,
                           dict(sheet=data_sheet,
                                title=title,
                                type=graph_obj['type'],
                                x=dict(title='Time', min_col=1, min_row=1, max_row=max_row),
                                y=dict(title=col_ref, min_col=2, max_col=max_col + 1, min_row=1, max_row=max_row)))
        t_content.append(dict(merge=4, font='link', align='wrap', hyper='#' + sname + '!A1', disp=col_ref))

        sheet_index += 1
        graph_num += 1

    return ml


def _graphs(switch_obj, single_port_graph_in, stats_graph_in, graph_type):
    """Parses the graphing information from the command line into a list of machine-readable dictionaries as follows:

    +-----------+---------------------------------------------------------------------------------------------------+
    | key       | Description                                                                                       |
    +===========+===================================================================================================+
    | stat      | Only present if -gs was entered on the command line. This is the fibrechannel-statistics to plot. |
    +-----------+---------------------------------------------------------------------------------------------------+
    | type      | Graph type. See brcddb.report.graph.chart_types.                                                  |
    +-----------+---------------------------------------------------------------------------------------------------+
    | port      | Only present if -gp was entered on the command line. This is the port number in s/p notation to   |
    |           | plot.                                                                                             |
    +-----------+---------------------------------------------------------------------------------------------------+
    | params    | If stat is not None, this is the list of ports whose statistic is to be plotted. If port is not   |
    |           | None, this is the list of statistics for the port to be plotted.                                  |
    +-----------+---------------------------------------------------------------------------------------------------+

    :param switch_obj: First switch object with list of ports
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param single_port_graph_in: Command line text for single port graphs. See help for -gp option
    :type single_port_graph_in: str
    :param stats_graph_in: Command line text for statistics graphs. See help for -gs option
    :type stats_graph_in: str
    :param graph_type: Type of graph
    :type graph_type: str
    :return graphs: List of dictionaries that define the graphs. See description of graph in _write_report() for details
    :rtype graphs: list
    :return messages: Error and warning messages
    :rtype messages: list
    """
    # Figure out the graph type and set up the return list of graphs
    ml = list()
    graphs = list()
    if graph_type is None:
        graph_type = 'line'
    elif graph_type not in report_graph.chart_types.keys():
        ml.append('Invalid graph type:   ' + graph_type + '. Defaulting to line')
        graph_type = 'line'

    # Single port, multiple statistics
    if isinstance(single_port_graph_in, str):
        for buf in single_port_graph_in.split(';'):
            temp_l = buf.split(',')
            if len(temp_l) > 1:
                port = temp_l.pop(0)
                if '/' not in port:
                    port = '0/' + port
                if switch_obj.r_port_obj(port) is None:
                    ml.append(port + ' not found. Skipping this port')
                else:
                    graphs.append(dict(port=port, parms=temp_l, type=graph_type))

    # statistic, multiple ports
    if isinstance(stats_graph_in, str):
        for buf in stats_graph_in.split(';'):
            temp_l = buf.split(',')
            if len(temp_l) > 1:
                statistic = 'fibrechannel-statistics/' + temp_l[0]
                to_graph = dict(stat=statistic, type=graph_type)
                temp_l = temp_l[1:]

                # If top or avg was specified for the ports, figure out the top (peak) ports for this statistic and the
                # top ports for the maximum sum for this statistic
                if 'top-' in temp_l[0].lower() or 'avg-' in temp_l[0].lower():
                    n = int(temp_l[0].split('-')[1])
                    switch_obj_l = [o for o in switch_obj.r_project_obj().r_switch_objects() if '-' in o.r_obj_key()]
                    port_total_d, port_peak_d, port_obj_l = dict(), dict(), list()
                    for port_obj in switch_obj_l.pop(0).r_port_objects():
                        port_stat = port_obj.r_get(statistic)
                        if port_stat is not None:
                            port_total_d.update({port_obj.r_obj_key(): port_stat})
                            port_peak_d.update({port_obj.r_obj_key(): port_stat})
                            port_obj_l.append(port_obj)
                    for switch_obj in switch_obj_l:
                        for port_obj in switch_obj.r_port_objects():
                            port_stat = port_obj.r_get(statistic)
                            if port_stat is not None:
                                port_key = port_obj.r_obj_key()
                                if port_key in port_total_d:
                                    port_total_d[port_key] += port_stat
                                    if port_stat > port_peak_d[port_key]:
                                        port_peak_d[port_key] = port_stat
                    for port_obj in port_obj_l:
                        port_obj.s_new_key('_peak', port_peak_d[port_obj.r_obj_key()], True)
                        port_obj.s_new_key('_total', port_total_d[port_obj.r_obj_key()], True)
                    peak_ports = gen_util.sort_obj_num(port_obj_l, '_peak', True)[0: min(n, len(port_obj_l))]
                    max_ports = gen_util.sort_obj_num(port_obj_l, '_total', True)[0: min(n, len(port_obj_l))]

                    # Above sorts by port object. All we want is the port number
                    if 'top' in temp_l[0]:
                        to_graph.update(parms=[port_obj.r_obj_key() for port_obj in peak_ports])
                    else:
                        to_graph.update(parms=[port_obj.r_obj_key() for port_obj in max_ports])

                elif 'eport' in temp_l[0].lower().replace('-', ''):
                    port_list = brcddb_search.match_test(switch_obj.r_port_objects, brcddb_search.e_ports)
                    if len(port_list) == 0:
                        ml.append('No E-Ports found')
                    to_graph.update(parms=[port_obj.r_obj_key() for port_obj in port_list])

                else:  # It's a list of ports. Make sure they are valid and prepend '0/' if necessary
                    port_list = list()
                    for port in temp_l:
                        mod_port = port if '/' in port else '0/' + port
                        if switch_obj.r_port_obj(mod_port) is None:
                            ml.append('Invalid port number or port not found in switch: ' + mod_port)
                        else:
                            port_list.append(mod_port)
                    to_graph.update(parms=port_list)

                if len(to_graph['parms']) > 0:
                    graphs.append(to_graph)

            else:
                ml.append('Missing parameter in ' + buf)

    return graphs, ml


def _write_report(switch_obj, report, graph_list, ml):
    """Creates an Excel workbook with the port statistics differences. Operates off global data

    :param switch_obj: Base switch object with ports to write to report
    :type switch_obj: brcddb.classes.switch.SwitchObj
    :param report: Name of report (Excel file name)
    :type report: str
    :param graph_list: List of dictionaries as returned from _graphs
    :type graph_list: list
    :return: Status code. See brcddb.brcddb_common.EXIT_xxxx
    :rtype: int
    """

    # Get the project and set up the workbook
    brcdapi_log.log('Generating Report: ' + report + '. This may take several seconds', echo=True)
    proj_obj = switch_obj.r_project_obj()
    wb = excel_util.new_report()

    # Set up the Project summary sheet with table of content
    title = 'Port Performance'
    tc_page = 'Project_Summary'
    t_content = [
        dict(new_row=False, merge=2, font='std', align='wrap', disp='Description'),
        dict(merge=2, font='std', align='wrap', disp=proj_obj.c_description()),
        dict(merge=2, new_row=False, font='std', align='wrap', disp='Data collected'),
        dict(merge=2, font='std', align='wrap', disp=proj_obj.r_date()),
    ]
    t_content_p = [
        dict(merge=4, font='std', align='wrap', disp=''),
        dict(merge=4, font='hdr_2', align='wrap', disp='Collected Port Data'),
        dict(merge=4, font='std', align='wrap', disp=''),
        dict(new_row=False, font='hdr_2', align='wrap', disp='Port'),
        dict(new_row=False, font='hdr_2', align='wrap', disp='Name'),
        dict(new_row=False, font='hdr_2', align='wrap', disp='Type'),
        dict(new_row=False, font='hdr_2', align='wrap', disp='State'),
        dict(font='hdr_2', align='wrap', disp='Description'),
    ]
    t_content_g = [
        dict(merge=4, font='std', align='wrap', disp=''),
        dict(merge=4, font='hdr_2', align='wrap', disp='Graphs'),
        dict(merge=4, font='std', align='wrap', disp=''),
    ]

    # Add the individual graphs and port sheets
    _add_ports(wb, tc_page, t_content_p, 0, switch_obj)
    ml.extend(_add_graphs(wb, tc_page, t_content_g, 0, switch_obj, graph_list))
    ml = gen_util.remove_duplicates(ml)
    t_content.extend(t_content_g + t_content_p)

    # Add the table of contents and save the report.
    report_utils.title_page(wb, None, tc_page, 0, title, t_content, (12, 22, 16, 10, 64))
    ml.append('Saving the report.')
    brcdapi_log.log(ml, echo=True)
    excel_util.save_report(wb, report)

    return brcddb_common.EXIT_STATUS_OK


def pseudo_main(report, in_f, single_port_graph, stats_graph, graph_type):
    """Basically the main(). Did it this way, so it can easily be used as a standalone module or called from another.

    :param report: Name of Excel report.
    :type report: str
    :param in_f: Name of data input file. This must be the output file, -o, from stats_c.py.
    :type in_f: str
    :param single_port_graph: Name of the worksheet to create with a graph of one or more stats counters for a port.
    :type single_port_graph: None, str
    :param stats_graph: Name of the worksheet to create with a graph of one or more statistical counters for a port.
    :type stats_graph: None, str
    :param graph_type: Type of graph. see report_graph.chart_types for details
    :type graph_type: str, None
    :return: Exit code. See exit codes in brcddb.brcddb_common
    :rtype: int
    """
    global __version__, _input_d

    # Read in the previously collected data
    brcdapi_log.log('Reading ' + in_f, echo=True)
    obj = brcdapi_file.read_dump(in_f)
    if obj is None:
        return brcddb_common.EXIT_STATUS_ERROR
    proj_obj = brcddb_project.new(obj.get('_obj_key'), obj.get('_date'))
    proj_obj.s_python_version(sys.version)
    proj_obj.s_description(obj.get('_description'))
    brcddb_copy.plain_copy_to_brcddb(obj, proj_obj)
    obj.clear()
    base_switch_obj = proj_obj.r_switch_obj(proj_obj.r_get('base_switch_wwn'))

    # Build the cross-reference tables.
    brcddb_util.build_login_port_map(proj_obj)  # Correlates name server logins with ports
    fab_obj = base_switch_obj.r_fabric_obj()
    if fab_obj is not None:
        brcddb_fabric.zone_analysis(base_switch_obj.r_fabric_obj())  # Determines what zones each login participates in

    graph_list, msg_list = _graphs(base_switch_obj, single_port_graph, stats_graph, graph_type)
    brcdapi_log.log('Writing ' + report)
    return _write_report(base_switch_obj, report, graph_list, msg_list)


def _get_input():
    """Parses the module load command line

    :return: Exit code. See exit codes in brcddb.brcddb_common
    :rtype: int
    """
    global __version__, _input_d

    # Get command line input
    buf = 'Create Excel Workbook from statistics gathered with stats_c.py with optional graphing capabilities.'
    args_d = gen_util.get_input(buf, _input_d)

    # Set up logging
    brcdapi_log.open_log(
        folder=args_d['log'],
        suppress=args_d['sup'],
        no_log=args_d['nl'],
        version_d=brcdapi_util.get_import_modules()
    )

    # Command line feedback
    ml = [
        os.path.basename(__file__) + ' version: ' + __version__,
        'Report, -r:      ' + args_d['r'],
        'Input file, -i:  ' + args_d['i'],
        'Port graph, -gp: ' + str(args_d['gp']),
        'Stat graph, -gs: ' + str(args_d['gs']),
        'Graph type, -gt: ' + str(args_d['gt']),
        'Log, -log:       ' + str(args_d['log']),
        'No log, -nl:     ' + str(args_d['nl']),
        'Suppress, -sup:  ' + str(args_d['sup']),
        '',
        ]
    brcdapi_log.log(ml, echo=True)

    return pseudo_main(brcdapi_file.full_file_name(args_d['r'], '.xlsx'),
                       brcdapi_file.full_file_name(args_d['i'], '.json'),
                       args_d['gp'],
                       args_d['gs'],
                       args_d['gt'])


##################################################################
#
#                    Main Entry Point
#
###################################################################
if _DOC_STRING:
    print('_DOC_STRING is True. No processing')
    exit(0)

if _STAND_ALONE:
    _ec = _get_input()
    brcdapi_log.close_log(['', 'Processing Complete. Exit code: ' + str(_ec)], echo=True)
    exit(_ec)
