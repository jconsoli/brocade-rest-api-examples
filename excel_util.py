"""
Copyright 2023, 2024 Consoli Solutions, LLC.  All rights reserved.

Licensed under the Apache License, Version 2.0 (the "License"); you may not use this file except in compliance with
the License. You may also obtain a copy of the License at https://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software distributed under the License is distributed on an
"AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied. See the License for the specific
language governing permissions and limitations under the License.

The license is free for single customer use (internal applications). Use of this module in the production,
redistribution, or service delivery for commerce requires an additional license. Contact jack@consoli-solutions.com for
details.

:mod:`brcdapi.excel_util` - Contains miscellaneous Excel workbook utilitarian methods.

**Public Methods & Data**

+-----------------------+-------------------------------------------------------------------------------------------+
| Method                | Description                                                                               |
+=======================+===========================================================================================+
| cell_match_val        | Finds the cell matching a value                                                           |
+-----------------------+-------------------------------------------------------------------------------------------+
| cell_update           | A convenient way to set cell properties and the cell value in a single call.              |
+-----------------------+-------------------------------------------------------------------------------------------+
| col_to_num            | Converts a cell reference to a column number. I'm pretty sure the openpyxl library has an |
|                       | equivalent. I couldn't find it so this was an expedient                                   |
+-----------------------+-------------------------------------------------------------------------------------------+
| copy_worksheet        | Typically used to copy a worksheet from one workbook to another                           |
+-----------------------+-------------------------------------------------------------------------------------------+
| excel_datetime        | Converts a datetime.datetime class object from Excel to formatted text                    |
+-----------------------+-------------------------------------------------------------------------------------------+
| find_headers          | Match columns to headers. Duplicate headers are ignored. Optionally warn if a duplicate   |
|                       | is encountered.                                                                           |
+-----------------------+-------------------------------------------------------------------------------------------+
| new_report            | Creates a workbook object for the Excel report.                                           |
+-----------------------+-------------------------------------------------------------------------------------------+
| parse_parameters      | Parses a Workbook into a dictionary of header columns and content by header. See          |
|                       | sample_parameters.xlsx                                                                    |
+-----------------------+-------------------------------------------------------------------------------------------+
| read_sheet            | Reads the contents (values) of a worksheet into two lists of dictionaries. The cell list  |
|                       | is a list of dictionaries whereby the key is the cell reference and the value is the      |
|                       | value of the cell. This is typically only used by applications using the brcddb search    |
|                       | engine. The second is a list of lists that make up a C like array that can be accessed    |
|                       | with a row and column number for the cell value.                                          |
+-----------------------+-------------------------------------------------------------------------------------------+
| read_workbook         | Reads a workbook into a list of worksheets followed by lists of lists which effectively   |
|                       | make up a row by column matrix of each sheet.                                             |
+-----------------------+-------------------------------------------------------------------------------------------+
| save_report           | Saves a workbook object as an Excel file.                                                 |
+-----------------------+-------------------------------------------------------------------------------------------+

**Version Control**

+-----------+---------------+---------------------------------------------------------------------------------------+
| Version   | Last Edit     | Description                                                                           |
+===========+===============+=======================================================================================+
| 4.0.0     | 04 Aug 2023   | Re-Launch                                                                             |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.1     | 06 Mar 2024   | Documentation updates only.                                                           |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.2     | 03 Apr 2024   | Added creator to save_report()                                                        |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.3     | 15 May 2024   | Added hidden parameter to read_sheet() and read_workbook()                            |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.4     | 20 Oct 2024   | Added comments and conditional formatting, cf, to cell_update()                       |
+-----------+---------------+---------------------------------------------------------------------------------------+
| 4.0.5     | 06 Dec 2024   | Added hidden parameter to copy_worksheet()                                            |
+-----------+---------------+---------------------------------------------------------------------------------------+
"""
__author__ = 'Jack Consoli'
__copyright__ = 'Copyright 2023, 2024 Consoli Solutions, LLC'
__date__ = '06 Dec 2024'
__license__ = 'Apache License, Version 2.0'
__email__ = 'jack@consoli-solutions.com'
__maintainer__ = 'Jack Consoli'
__status__ = 'Released'
__version__ = '4.0.5'

import openpyxl as xl
import openpyxl.utils.cell as xl_util
from openpyxl.comments import Comment
from openpyxl import Workbook
import fnmatch
import re
import os
import brcdapi.log as brcdapi_log
import brcdapi.file as brcdapi_file
import brcdapi.gen_util as gen_util

_DEFAULT_COMMENT_WIDTH = 400
_DEFAULT_COMMENT_HEIGHT = 100

# Use this to create a sheet name that is not only valid for Excel but can have a link. Note when creating a link to a
# sheet in Excel, there are additional restrictions on the sheet name. For example, it cannot contain a space. Sample
# use: good_sheet_name = valid_sheet_name.sub('_', bad_sheet_name)
valid_sheet_name = re.compile(r'[^\d\w_]')

# Using excel_datetime is clumsy. This is easier. Not that I need speed, but it's also faster. Used in excel_datetime
_num_to_month = ('Inv', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec')


class Found(Exception):
    pass


def excel_datetime(v, granularity):
    """Converts a datetime.datetime class object from Excel to formatted text

    :param v: Cell value of class type datetime.datetime
    :type v: datetime
    :param granularity: datetime conversion granularity 0: yyyy, 1: mm yyyy, 2: dd mm yyyy, 3: dd mm yyyy hh, \
        4: dd mm yyyy hh:mm, 5: dd mm yyyy hh:mm:ss, 6: dd mm yyyy hh:mm:ss:uuu
    :type granularity: int
    """
    global _num_to_month

    msec = str(v.microsecond)
    tl = [str(v.year),
          _num_to_month[v.month] + ' ',
          '0' + str(v.day) + ' ' if v.day < 10 else str(v.day) + ' ',
          ' 0' + str(v.hour) if v.hour < 10 else ' ' + str(v.hour),
          ':0' + str(v.minute) if v.minute < 10 else ':' + str(v.minute),
          ':0' + str(v.second) if v.second < 10 else ':' + str(v.second),
          ':00' + msec if len(msec) == 1 else ':0' + msec if len(msec) == 2 else ':' + msec[0:3]]
    buf = ''
    for i in range(0, min(granularity+1, len(tl))):
        buf = buf + tl[i] if i > 2 else tl[i] + buf

    return buf


def parse_parameters(in_wb=None, sheet_name='parameters', hdr_row=0, wb_name=None):
    """Parses a Workbook into a dictionary of header columns and content by header. See sample_parameters.xlsx

    Returned dictionary is as follows:

    +-----------+-------+-------------------------------------------------------------------------------------------+
    | key       | Value | Value                                                                                     |
    |           | type  |                                                                                           |
    +===========+=======+===========================================================================================+
    | hdr_col   | dict  | key: Header, value: 0 based column number                                                 |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | content   | list  | A list of dictionaries. For each dictionary, the key is the header and the value is the   |
    |           |       | matching cell value. The list order is in row order.                                      |
    +-----------+-------+-------------------------------------------------------------------------------------------+
    | col_width | list  | List of column widths (float). Index into the list is the zero relative column number.    |
    |           |       | For example, the width of column A is col_width[0]. WARNING: this is taken from           |
    |           |       | column_dimensions() which doesn't always return a column dimension so the entry in this   |
    |           |       | list may be None.                                                                         |
    +-----------+-------+-------------------------------------------------------------------------------------------+

    :param in_wb: Workbook object returned from openpyxl.load_workbook() to read if wb_name is None
    :type in_wb: openpyxl.workbook.workbook.Workbook, None
    :param sheet_name: Name of sheet to read
    :type sheet_name: str
    :param hdr_row: Header row. Rows below are assumed to be the content data
    :type hdr_row: int
    :param wb_name: Name of workbook to read. If None, assume in_wb is a valid Workbook object
    :type wb_name: str, None
    :return: Dictionary as described above
    :rtype: dict
    """
    col_width, hdr_d, content_l = list(), dict(), list()
    rd = dict(col_width=col_width, hdr_col=hdr_d, content=content_l)  # Return dict
    wb = in_wb if wb_name is None else xl.load_workbook(wb_name, data_only=True)

    # Figure out what and where the headers are
    try:
        sheet = wb[sheet_name]
    except BaseException as e:
        e_buf = str(type(e)) + ': ' + str(e)
        brcdapi_log.exception(['sheet ' + sheet_name + ' does not exist.', 'Exception is: ' + e_buf], echo=True)
        return rd
    sl, al = read_sheet(sheet, 'row')
    hdr_row_l = al[hdr_row]
    for i in range(0, len(hdr_row_l)):
        col_width.append(sheet.column_dimensions[xl_util.get_column_letter(i+1)].width)
        if hdr_row_l[i] is not None:
            hdr_d.update({hdr_row_l[i]: i})

    # Read in the values for each column
    for i in range(hdr_row+1, len(al)):
        d = dict()
        for key, col in hdr_d.items():
            d.update({key: al[i][col]})
        content_l.append(d)

    return rd


def new_report():
    """Creates a workbook object for the Excel report.

    :return: wb
    :rtype: Workbook object
    """
    return xl.Workbook()


def save_report(wb, file_name='Report.xlsx', creator='Consoli-Solutions, LLC'):
    """Saves a workbook object as an Excel file.

    :param wb: Workbook object
    :type wb: openpyxl.workbook.workbook.Workbook
    :param file_name: Report name
    :type file_name: str
    :param creator: Name of person or organization creating this document.
    :type creator: str
    """
    wb.properties.creator = creator
    wb.save(brcdapi_file.full_file_name(file_name, '.xlsx'))


def col_to_num(cell):
    """Converts a cell reference to a column number.

    :param cell: Excel spreadsheet cell reference. Example: 'AR20' or just 'AR'
    :type cell: str
    :return: Column number. 0 if column not found
    :rtype: int
    """
    r = 0
    for i in range(0, len(cell)):
        x = ord(cell[i].upper()) - 64
        if x < 1 or x > 26:
            break
        r = (r * 26) + x  # Python should understand algebraic hierarchy, but I'm not leaving anything to chance.

    return r


def cell_match_val(sheet, val, col=None, row=None, num=1):
    """Finds the cell matching a value

    :param sheet: Sheet structure returned from wb.create_sheet()
    :type sheet: class
    :param val: Cell contents we're looking for
    :type val: int, float, str
    :param col: List of columns letters to look in. If None, checks all columns on sheet.
    :type col: list, str, None
    :param row: Row number or list of row numbers to look in. If None, checks all rows on sheet.
    :type row: int, list, None
    :param num: Number of instances to find
    :type num: int
    :return: List of cell references where value found. If num == 1: just one str is returned. None if not found
    :rtype: str, list, None
    """
    col_list = [xl_util.get_column_letter(i) for i in range(1, sheet.max_column)] if col is None \
        else gen_util.convert_to_list(col)
    row_list = [i for i in range(1, sheet.max_row)] if row is None else gen_util.convert_to_list(row)

    ret = list()
    try:
        for c in col_list:
            for r in row_list:
                cell = c + str(r)
                rv = sheet[cell].value
                if (isinstance(val, (int, float)) and isinstance(rv, (int, float))) or \
                        (isinstance(val, str) and isinstance(rv, str)):
                    if val == rv:
                        ret.append(cell)
                        if num >= len(ret):
                            raise Found
    except Found:
        pass

    return ret if num != 1 else ret[0] if len(ret) > 0 else None


def read_sheet(sheet, order='col', granularity=2, hidden=True):
    """Reads the contents (values) of a worksheet into two lists of dictionaries. The cell list is a list of
    dictionaries whereby the key is the cell reference and the value is the value of the cell. This is typically only
    used by applications using the brcddb search engine. The second is a list of lists that make up a C like array that
    can be accessed with a row and column number for the cell value.

    sl Detail:
    +---------------+---------------------------------------------------------------------------------------+
    | key           | Value description                                                                     |
    +===============+=======================================================================================+
    | cell          | Cell reference.                                                                       |
    +---------------+---------------------------------------------------------------------------------------+
    | val           | Value read from cell. Special types (not int, float, or str) are converted to None    |
    +---------------+---------------------------------------------------------------------------------------+

    Intended to be used by methods that will feed this list to brcddb.utils.search.match_test()

    :param sheet: Sheet structure returned from wb.create_sheet()
    :type sheet: class
    :param order: Order in which to read. 'row' means read by row, then each individual column. 'col' for column 1st
    :type order: str
    :param granularity: See description of granularity with excel_datetime()
    :type granularity: int
    :param hidden: If True (default), read all cells from hidden rows/columns. If False, all values for cells in hidden
                   rows/columns will be None.
    :type hidden: bool
    :return sl: Dictionaries as noted above
    :rtype sl: list
    :return al: Contents of the worksheet referenced by al[col-1][row-1] if order is 'col' or
                al[row-1][col-1] if order is 'row'
    :rtype al: list
    """
    # Read in all the cell values
    sl, al = list(), list()
    if order.lower() == 'col':
        for col in range(1, sheet.max_column+1):
            if not hidden and col in sheet.column_dimensions and sheet.column_dimensions[col].hidden:
                al.append([None for row in range(1, sheet.max_row+1)])
                continue
            col_ref = xl_util.get_column_letter(col)
            rl = list()
            for row in range(1, sheet.max_row+1):
                cell = col_ref + str(row)
                v = sheet[cell].value
                if isinstance(v, (bool, int, float, str)):
                    sl.append(dict(cell=cell, val=v))
                    rl.append(v)
                elif 'datetime.datetime' in str(type(v)):
                    buf = excel_datetime(v, granularity)
                    sl.append(dict(cell=cell, val=buf))
                    rl.append(buf)
                else:
                    rl.append(None)
            al.append(rl)
    else:
        for row in range(1, sheet.max_row+1):
            if not hidden and row in sheet.row_dimensions and sheet.row_dimensions[row].hidden:
                al.append([None for col in range(1, sheet.max_column + 1)])
                continue
            cl = list()
            for col in range(1, sheet.max_column+1):
                cell = xl_util.get_column_letter(col) + str(row)
                sheet_cell = sheet[cell]
                v = sheet_cell.value
                if isinstance(v, (bool, int, float, str)):
                    sl.append(dict(cell=cell, val=v))
                    cl.append(v)
                elif 'datetime.datetime' in str(type(v)):
                    buf = excel_datetime(v, granularity)
                    sl.append(dict(cell=cell, val=buf))
                    cl.append(buf)
                else:
                    cl.append(None)
            al.append(cl)

    return sl, al


def cell_update(sheet, row, col, buf, font=None, align=None, fill=None, link=None, border=None, comments=None, cf=None,
                dv=None, number_format=None, comment_width=_DEFAULT_COMMENT_WIDTH, comment_height=_DEFAULT_COMMENT_HEIGHT):
    """A convenient way to set cell properties and the cell value in a single call.

    :param sheet: openpyxl worksheet
    :type sheet: Worksheet
    :param row: Sheet row number
    :type row: int
    :param col: Sheet column by number
    :type col: int
    :param buf: Cell content
    :type buf: None, str, int, float
    :param font: Font type
    :type font: None, xl_styles
    :param align: Cell alignment
    :type align: None, xl_styles
    :param fill: Cell fill color
    :type fill: None, xl_styles
    :param link: Hyperlink to apply to cell
    :type link: None, xl_styles
    :param border: Border type to apply to cell
    :type border: None, xl_styles
    :param comments: Cell comments
    :type comments: None, str, list
    :param cf: Conditional formatting rule
    :type cf: None, openpyxl.formatting.rule.Rule
    :param dv: Data validation
    :type dv: None, openpyxl.worksheet.datavalidation.DataValidation
    :param number_format: See openpyxl number_format for details
    :type number_format: None, str
    :param comment_width: Width of comment pop-up. The default is _DEFAULT_COMMENT_WIDTH
    :type comment_width: int
    :param comment_height: Height of comment pop-up. The default is _DEFAULT_COMMENT_HEIGHT
    :type comment_height: int
    :return: None
    :rtype: None
    """
    cell = xl_util.get_column_letter(col) + str(row)
    if font is not None:
        sheet[cell].font = font
    if fill is not None:
        sheet[cell].fill = fill
    if align is not None:
        sheet[cell].alignment = align
    if border is not None:
        sheet[cell].border = border
    if link is not None:
        sheet[cell].hyperlink = link
    if buf is not None:
        sheet[cell] = buf
    if comments is not None:
        comment = Comment(comments, 'Consoli Solutions, LLC')
        comment.width = comment_width
        comment.height = comment_height
        sheet[cell].comment = comment
    if cf is not None:
        sheet.conditional_formatting.add(cell, cf)
    if dv is not None:
        dv.add(cell)
    if number_format is not None:
        sheet[cell].number_format = number_format


def read_workbook(file, dm=0, order='row', sheets=None, skip_sheets=None, echo=False, hidden=True):
    """Reads an Excel workbook

    For large workbooks that take a long time to read, it turned out to be convenient to leave these debug modes in.
    Note that reading a workbook is very time-consuming while reading a JSON file is magnitudes of order faster.

    +-------+-------------------------------------------------------------------------------------------------------|
    | dm    | Description                                                                                           |
    +=======+=======================================================================================================|
    | 0     | Read the workbook normally                                                                            |
    +-------+-------------------------------------------------------------------------------------------------------|
    | 1     | Read the workbook then write the return list to a JSON file of the same name replacing '.xlsx' with   |
    |       | '.json'                                                                                               |
    +-------+-------------------------------------------------------------------------------------------------------|
    | 2     | Replace .xlsx in the file name with .json, then read the JSON file. sheets and skip_sheets are        |
    |       | ignored.                                                                                              |
    +-------+-------------------------------------------------------------------------------------------------------|
    | 3     | If the equivalent JSON file exists and the last modification time stamp is more recent than the Excel |
    |       | file continue as mode 2. Otherwise, continue as mode 1.                                               |
    +-------+-------------------------------------------------------------------------------------------------------|

    :param file: Name of Excel workbook to read
    :type file: str
    :param dm: Debug mode.
    :type dm: int
    :param order: 'row' for row first followed by column data. 'col' for columns first followed by rows
    :type order: str
    :param sheets: Sheet or list of sheets by name to read. If None, read all not in skip_sheets. Accepts wild cards
    :type sheets: None, list, tuple, str
    :param skip_sheets: Sheet or list of sheets to skip. Accepts wild cards.
    :type skip_sheets: None, list, tuple, str
    :param echo: If True, print read/write status to STD_OUT
    :type echo: bool
    :param hidden: If True, read hidden rows from sheet
    :type hidden: bool
    :return el: Errors. Empty if no errors.
    :rtype el: list
    :return sl: List of dictionaries, one for each sheet, with the file, sheet name, and excel_util.read_sheet() output.
    :rtype sl: list
    """
    el, rl = list(), list()
    json_file = file.replace('.xlsx', '.json')

    if dm >= 2:
        # Try reading the JSON file
        try:
            excel_file_time = os.path.getmtime(file)
        except (FileExistsError, FileNotFoundError):
            excel_file_time = 0
        try:
            if excel_file_time < os.path.getmtime(json_file):
                brcdapi_log.log('Reading: ' + json_file, echo=echo)
                return el, brcdapi_file.read_dump(json_file)
        except FileNotFoundError:
            pass
        except FileExistsError:
            el.append('The folder in ' + file + ' does not exist.')
            return el, rl

    # Read the workbook
    brcdapi_log.log('Reading ' + file, echo=echo)
    try:
        wb = xl.load_workbook(file, data_only=True)  # Read the Workbook
    except FileNotFoundError:
        el.append('File not found: ' + file)
        return el, rl
    except FileExistsError:
        el('The folder in ' + file + ' does not exist.')
        return el, rl
    except TypeError:
        buf = 'Encountered a TypeError reading ' + file + '. This typically occurs when there is a sheet name '
        buf += 'with special characters. To fix this, rename the sheet name (sheet tab).'
        el.append(buf)
        return el, rl

    # Figure out which sheets to skip
    skip_sheet_d = dict()
    for sheet_name in gen_util.convert_to_list(skip_sheets):
        if '*' in sheet_name or '?' in sheet_name:
            for x_sheet_name in fnmatch.filter(wb.sheetnames, sheet_name):
                skip_sheet_d.update({x_sheet_name: True})
        else:
            skip_sheet_d.update({sheet_name: True})

    # Figure out which sheets to read
    sheet_l = list()
    if sheets is None:
        sheet_l.extend([buf for buf in wb.sheetnames if buf not in skip_sheet_d])
    else:
        for sheet_name in gen_util.convert_to_list(sheets):
            if '*' in sheet_name or '?' in sheet_name:
                for buf in fnmatch.filter(wb.sheetnames, sheet_name):
                    if buf not in skip_sheet_d and buf not in sheet_l:
                        sheet_l.append(buf)
            elif sheet_name in wb.sheetnames:
                if sheet_name not in skip_sheet_d and sheet_name not in sheet_l:
                    sheet_l.append(sheet_name)
            else:
                brcdapi_log.log('Sheet ' + sheet_name + ' not found in ' + file + '. Skipping this sheet.', echo=True)

    # Read the sheets
    for sheet_name in sheet_l:
        brcdapi_log.log('  Reading sheet ' + sheet_name, echo=echo)
        sl, al = read_sheet(wb[sheet_name], order='row', hidden=hidden)
        rl.append(dict(file=file, sheet=sheet_name, sl=sl, al=al))
    brcdapi_log.log('  Read complete', echo=echo)

    if dm == 1 or dm == 3:
        # Write out to JSON
        brcdapi_log.log('  Writing ' + json_file, echo=echo)
        brcdapi_file.write_dump(rl, json_file)
        brcdapi_log.log('  Write complete', echo=echo)

    return el, rl


def copy_worksheet(wb, sheet_index, sheet_name, sheet_l, col_width_l=None, font=None, align=None, fill=None,
                   border=None, hidden=False):
    """Typically used to copy a worksheet from one workbook to another

    :param wb: openpyxl Workbook
    :type wb: openpyxl class object
    :param sheet_index: Index as to where this sheet should be placed in the Workbook
    :type sheet_index: int
    :param sheet_name: Name of worksheet
    :type sheet_name: str
    :param sheet_l: Sheets list returned from excel_util.read_workbook
    :type sheet_l: list
    :param col_width_l: Column widths. First entry is for 'A', next is for 'B', etc.
    :type col_width_l: list, tuple, int, None
    :param font: Font type
    :type font: None, xl_styles
    :param align: Cell alignment
    :type align: None, xl_styles
    :param fill: Cell fill color
    :type fill: None, xl_styles
    :param border: Border type to apply to cell
    :type border: None, xl_styles
    :param hidden: If True, set the sheet state to hidden
    :type hidden: bool
    :return: List of error messages
    :rtype: list
    """
    el, sheet_d = list(), dict()

    # Find the sheet to copy
    for sheet_d in sheet_l:
        if str(sheet_d.get('sheet')) == sheet_name:

            # Create the sheet and set up the column widths
            sheet = wb.create_sheet(index=sheet_index, title=sheet_name)
            if hidden:
                sheet.sheet_state = 'hidden'
            col = 1
            for width in gen_util.convert_to_list(col_width_l):
                sheet.column_dimensions[xl_util.get_column_letter(col)].width = width
                col += 1

            # Copy the data
            for cell_d in sheet_d['sl']:
                cell = cell_d['cell']
                if font is not None:
                    sheet[cell].font = font
                if fill is not None:
                    sheet[cell].fill = fill
                if align is not None:
                    sheet[cell].alignment = align
                if border is not None:
                    sheet[cell].border = border
                sheet[cell] = cell_d['val']
            return el

    # If we got this far, the sheet wasn't found
    el.append('Could not find ' + sheet_name + ' in ' + str(sheet_d.get('file')))

    return el


def find_headers(hdr_row_l, hdr_l=None, warn=False, plus_one=False):
    """Match columns to headers. Duplicate headers are ignored. Optionally warn if a duplicate is encountered.

    :param hdr_row_l: Typically, al[0] from sl, al = excel_util.read_sheet(sheet, 'row')
    :type hdr_row_l: list
    :param hdr_l: Header or list of headers to find. Find all headers if None
    :type hdr_l: str, list, tuple, None
    :param warn: If True, add an exception message to the log warning that there are multiple column headers
    :type warn: bool
    :param plus_one: Lists begin with index zero. Excel columns begin with 1. If True, returns an Excel column
    :type plus_one: bool
    :return: Dictionary of headers. Key is the header in hdr_l. The value is the index into hdr_row where it was found.
             if not found, the value is None
    :rtype: dict, None
    """
    rd = dict()

    # If None, get all headers
    if hdr_l is None:
        for col in range(0, len(hdr_row_l)):
            if hdr_row_l[col] in rd:
                if warn:
                    brcdapi_log.exception('Duplicate header: ' + hdr_row_l[col], echo=True)
            else:
                rd.update({hdr_row_l[col]: col})

    # There is a specific list of headers to get
    else:
        for buf in gen_util.convert_to_list(hdr_l):
            rd.update({buf: None})
            for col in range(0, len(hdr_row_l)):
                if hdr_row_l[col] == buf:
                    rd[buf] = col
                    break

    # Adjust for Excel column numbering.
    if plus_one:
        for key in rd.keys():
            rd[key] += 1

    return rd
